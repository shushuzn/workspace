#!/usr/bin/env python3
import json, urllib.request, time, hashlib, re
from pathlib import Path
from datetime import datetime, timedelta
from functools import lru_cache
from collections import defaultdict

WORKSPACE = Path("D:/OpenClaw/workspace")
OUTPUT = WORKSPACE / "50-reports" / "stocks"
OUTPUT.mkdir(parents=True, exist_ok=True)

# Cache for performance
_cache = {}
_cache_ttl = 300  # 5 minutes

# Config file
CONFIG_FILE = WORKSPACE / "30-scripts-tools" / "stock_pro_config.json"

# News cache
NEWS_CACHE_FILE = OUTPUT / "news_cache.json"

def load_config():
    """Load configuration from file"""
    if CONFIG_FILE.exists():
        try:
            with open(CONFIG_FILE, 'r', encoding='utf-8') as f:
                return json.load(f)
        except Exception:
            pass
    return {
        "watchlist": ["NVDA", "META", "JPM", "AAPL", "MSFT", "GOOGL", "AMZN", "TSLA"],
        "alert_threshold": 30,
        "default_format": "text",
        "refresh_interval": 30,
        "notifications": {
            "enabled": False,
            "webhook_url": "",
            "email": ""
        }
    }

def save_config(config):
    """Save configuration to file"""
    with open(CONFIG_FILE, 'w', encoding='utf-8') as f:
        json.dump(config, f, indent=2)

def get_cached(key, fetch_func, ttl=_cache_ttl):
    """Get cached result or fetch new data"""
    now = time.time()
    if key in _cache:
        data, timestamp = _cache[key]
        if now - timestamp < ttl:
            return data
    data = fetch_func()
    _cache[key] = (data, now)
    return data

def clear_cache():
    """Clear all cache"""
    global _cache
    _cache = {}
    print("[Cache] Cleared")

def progress_bar(current, total, prefix='', length=40):
    """Show progress bar"""
    percent = current / total
    filled = int(length * percent)
    bar = '=' * filled + '-' * (length - filled)
    print(f'\r{prefix} [{bar}] {percent*100:.0f}%', end='', flush=True)
    if current >= total:
        print()

ANALYST = {
    # Technology
    "META":  {"target": 864, "rating": "Overweight", "num": 55},
    "NVDA":  {"target": 269, "rating": "Outperform", "num": 52},
    "AAPL":  {"target": 275, "rating": "Overweight", "num": 48},
    "MSFT":  {"target": 495, "rating": "Buy", "num": 54},
    "GOOGL": {"target": 365, "rating": "Buy", "num": 56},
    "AMZN":  {"target": 400, "rating": "Buy", "num": 55},
    "AMD":   {"target": 185, "rating": "Outperform", "num": 45},
    "TSLA":  {"target": 250, "rating": "Hold", "num": 40},
    "NFLX":  {"target": 700, "rating": "Hold", "num": 42},
    "CRM":   {"target": 320, "rating": "Buy", "num": 38},
    "ORCL":  {"target": 180, "rating": "Buy", "num": 35},
    "INTC":  {"target": 45, "rating": "Underperform", "num": 40},
    "QCOM":  {"target": 200, "rating": "Overweight", "num": 38},
    # Finance
    "JPM":   {"target": 370, "rating": "Overweight", "num": 42},
    "BAC":   {"target": 55, "rating": "Buy", "num": 38},
    "GS":    {"target": 650, "rating": "Buy", "num": 30},
    # Healthcare
    "JNJ":   {"target": 225, "rating": "Neutral", "num": 38},
    "UNH":   {"target": 600, "rating": "Buy", "num": 35},
    "PFE":   {"target": 32, "rating": "Hold", "num": 35},
    # Consumer
    "WMT":   {"target": 220, "rating": "Buy", "num": 35},
    "COST":  {"target": 950, "rating": "Buy", "num": 32},
    "KO":    {"target": 75, "rating": "Buy", "num": 30},
}

FD = {
    # Extended Financial Data: gm, nm, roe, roic, de, rev_g (revenue growth), fcf_yield, div_yield, payout, ev_ebitda
    "META":  {"gm": 0.81, "nm": 0.32, "roe": 0.25, "roic": 0.22, "de": 0.1, "rev_g": 0.25, "fcf_yield": 0.035, "div_yield": 0, "payout": 0, "ev_ebitda": 18},
    "NVDA":  {"gm": 0.73, "nm": 0.49, "roe": 0.58, "roic": 0.45, "de": 0.4, "rev_g": 0.50, "fcf_yield": 0.028, "div_yield": 0, "payout": 0, "ev_ebitda": 35},
    "AAPL":  {"gm": 0.47, "nm": 0.24, "roe": 0.45, "roic": 0.35, "de": 1.8, "rev_g": 0.04, "fcf_yield": 0.055, "div_yield": 0.0055, "payout": 0.15, "ev_ebitda": 22},
    "MSFT":  {"gm": 0.70, "nm": 0.35, "roe": 0.38, "roic": 0.28, "de": 0.5, "rev_g": 0.15, "fcf_yield": 0.030, "div_yield": 0.0075, "payout": 0.25, "ev_ebitda": 28},
    "GOOGL": {"gm": 0.56, "nm": 0.24, "roe": 0.28, "roic": 0.20, "de": 0.1, "rev_g": 0.12, "fcf_yield": 0.045, "div_yield": 0, "payout": 0, "ev_ebitda": 17},
    "AMZN":  {"gm": 0.48, "nm": 0.09, "roe": 0.22, "roic": 0.15, "de": 0.6, "rev_g": 0.11, "fcf_yield": 0.065, "div_yield": 0, "payout": 0, "ev_ebitda": 22},
    "JPM":   {"gm": 0.65, "nm": 0.34, "roe": 0.17, "roic": 0.12, "de": 1.3, "rev_g": 0.08, "fcf_yield": 0.085, "div_yield": 0.024, "payout": 0.30, "ev_ebitda": 10},
    "JNJ":   {"gm": 0.68, "nm": 0.20, "roe": 0.22, "roic": 0.16, "de": 0.5, "rev_g": 0.05, "fcf_yield": 0.060, "div_yield": 0.030, "payout": 0.45, "ev_ebitda": 14},
    "AMD":   {"gm": 0.47, "nm": 0.09, "roe": 0.18, "roic": 0.14, "de": 0.3, "rev_g": 0.20, "fcf_yield": 0.035, "div_yield": 0, "payout": 0, "ev_ebitda": 25},
    # Additional Stocks
    "TSLA":  {"gm": 0.18, "nm": 0.15, "roe": 0.25, "roic": 0.18, "de": 0.1, "rev_g": 0.20, "fcf_yield": 0.025, "div_yield": 0, "payout": 0, "ev_ebitda": 45},
    "NFLX":  {"gm": 0.42, "nm": 0.20, "roe": 0.30, "roic": 0.25, "de": 0.5, "rev_g": 0.15, "fcf_yield": 0.040, "div_yield": 0, "payout": 0, "ev_ebitda": 20},
    "CRM":   {"gm": 0.73, "nm": 0.12, "roe": 0.15, "roic": 0.10, "de": 0.3, "rev_g": 0.10, "fcf_yield": 0.035, "div_yield": 0, "payout": 0, "ev_ebitda": 30},
    "INTC":  {"gm": 0.40, "nm": 0.08, "roe": 0.10, "roic": 0.08, "de": 0.4, "rev_g": -0.02, "fcf_yield": 0.030, "div_yield": 0.015, "payout": 0.50, "ev_ebitda": 15},
    "BAC":   {"gm": 0.60, "nm": 0.28, "roe": 0.12, "roic": 0.08, "de": 1.0, "rev_g": 0.05, "fcf_yield": 0.075, "div_yield": 0.025, "payout": 0.30, "ev_ebitda": 12},
    "GS":    {"gm": 0.55, "nm": 0.25, "roe": 0.15, "roic": 0.10, "de": 4.0, "rev_g": 0.08, "fcf_yield": 0.060, "div_yield": 0.025, "payout": 0.25, "ev_ebitda": 14},
    "UNH":   {"gm": 0.24, "nm": 0.08, "roe": 0.22, "roic": 0.15, "de": 0.8, "rev_g": 0.08, "fcf_yield": 0.045, "div_yield": 0.015, "payout": 0.30, "ev_ebitda": 18},
    "WMT":   {"gm": 0.25, "nm": 0.03, "roe": 0.20, "roic": 0.15, "de": 0.6, "rev_g": 0.05, "fcf_yield": 0.055, "div_yield": 0.015, "payout": 0.40, "ev_ebitda": 12},
    "COST":  {"gm": 0.13, "nm": 0.03, "roe": 0.28, "roic": 0.18, "de": 0.3, "rev_g": 0.06, "fcf_yield": 0.035, "div_yield": 0.005, "payout": 0.25, "ev_ebitda": 20},
    "KO":    {"gm": 0.60, "nm": 0.22, "roe": 0.40, "roic": 0.12, "de": 1.7, "rev_g": 0.05, "fcf_yield": 0.050, "div_yield": 0.030, "payout": 0.70, "ev_ebitda": 16},
}

# Historical P/E percentiles (5-year)
PE_HIST = {
    "META":  {"curr": 35, "low": 15, "high": 45, "avg": 28},
    "NVDA":  {"curr": 35, "low": 15, "high": 120, "avg": 55},
    "AAPL":  {"curr": 32, "low": 20, "high": 40, "avg": 28},
    "MSFT":  {"curr": 36, "low": 22, "high": 45, "avg": 32},
    "GOOGL": {"curr": 25, "low": 18, "high": 35, "avg": 26},
    "AMZN":  {"curr": 45, "low": 30, "high": 120, "avg": 65},
    "JPM":   {"curr": 12, "low": 8, "high": 18, "avg": 12},
    "JNJ":   {"curr": 22, "low": 15, "high": 28, "avg": 20},
    "AMD":   {"curr": 28, "low": 10, "high": 80, "avg": 40},
    "TSLA":  {"curr": 65, "low": 30, "high": 200, "avg": 80},
    "NFLX":  {"curr": 35, "low": 20, "high": 60, "avg": 40},
    "CRM":   {"curr": 45, "low": 25, "high": 70, "avg": 50},
    "INTC":  {"curr": 25, "low": 10, "high": 35, "avg": 18},
    "BAC":   {"curr": 12, "low": 8, "high": 18, "avg": 12},
    "GS":    {"curr": 14, "low": 8, "high": 22, "avg": 14},
    "UNH":   {"curr": 22, "low": 18, "high": 35, "avg": 25},
    "WMT":   {"curr": 28, "low": 20, "high": 35, "avg": 28},
    "COST":  {"curr": 45, "low": 30, "high": 55, "avg": 40},
    "KO":    {"curr": 22, "low": 18, "high": 30, "avg": 24},
}

# Analyst ratings distribution (buy/hold/underperform percentages)
ANALYST_RATINGS = {
    "META":  {"buy": 75, "hold": 20, "sell": 5},
    "NVDA":  {"buy": 80, "hold": 15, "sell": 5},
    "AAPL":  {"buy": 60, "hold": 30, "sell": 10},
    "MSFT":  {"buy": 70, "hold": 25, "sell": 5},
    "GOOGL": {"buy": 65, "hold": 28, "sell": 7},
    "AMZN":  {"buy": 68, "hold": 25, "sell": 7},
    "JPM":   {"buy": 70, "hold": 25, "sell": 5},
    "JNJ":   {"buy": 45, "hold": 40, "sell": 15},
    "AMD":   {"buy": 70, "hold": 20, "sell": 10},
    "TSLA":  {"buy": 40, "hold": 35, "sell": 25},
    "NFLX":  {"buy": 50, "hold": 35, "sell": 15},
    "INTC":  {"buy": 30, "hold": 40, "sell": 30},
    "BAC":   {"buy": 55, "hold": 35, "sell": 10},
    "KO":    {"buy": 60, "hold": 35, "sell": 5},
}

def fetch_live(sym):
    """Fetch live price using yfinance"""
    try:
        import yfinance as yf
        ticker = yf.Ticker(sym)
        info = ticker.info

        price = info.get('currentPrice') or info.get('regularMarketPrice') or info.get('previousClose')
        mc = info.get('marketCap', 0)
        pe = info.get('trailingPE', 0) or 0
        beta = info.get('beta', 1.0) or 1.0
        eps_t = info.get('trailingEps', 0) or 0
        eps_f = info.get('forwardEps', 0) or 0
        peg = info.get('pegRatio', 0) or 2.0
        w52h = info.get('fiftyTwoWeekHigh', 0) or price * 1.2
        w52l = info.get('fiftyTwoWeekLow', 0) or price * 0.8

        return {
            "price": price or 0,
            "mc": mc,
            "pe": pe,
            "beta": beta,
            "eps_t": eps_t,
            "eps_f": eps_f,
            "peg": peg,
            "w52h": w52h,
            "w52l": w52l,
            "success": True,
            "source": "yfinance",
            "name": info.get('shortName', sym),
        }
    except Exception as e:
        return {"success": False, "error": str(e)}

def fetch(sym):
    # Fallback database for when API fails
    FALLBACK = {
        "META":  {"price": 593, "mc": 1.5e12, "pe": 35, "beta": 1.2, "eps_t": 23.5, "eps_f": 27.5, "peg": 1.4, "w52h": 740, "w52l": 350},
        "NVDA":  {"price": 172, "mc": 840e9, "pe": 35, "beta": 1.7, "eps_t": 4.90, "eps_f": 7.50, "peg": 0.9, "w52h": 974, "w52l": 138},
        "AAPL":  {"price": 248, "mc": 3.8e12, "pe": 32, "beta": 1.2, "eps_t": 7.75, "eps_f": 8.50, "peg": 3.0, "w52h": 315, "w52l": 164},
        "MSFT":  {"price": 382, "mc": 2.8e12, "pe": 36, "beta": 0.9, "eps_t": 10.5, "eps_f": 12.0, "peg": 2.5, "w52h": 468, "w52l": 275},
        "GOOGL": {"price": 301, "mc": 1.9e12, "pe": 25, "beta": 1.1, "eps_t": 6.5, "eps_f": 7.8, "peg": 1.2, "w52h": 420, "w52l": 275},
        "AMZN":  {"price": 205, "mc": 2.1e12, "pe": 45, "beta": 1.1, "eps_t": 4.55, "eps_f": 6.0, "peg": 2.0, "w52h": 360, "w52l": 185},
        "JPM":   {"price": 287, "mc": 415e9, "pe": 12, "beta": 1.1, "eps_t": 23.5, "eps_f": 25.5, "peg": 1.8, "w52h": 430, "w52l": 280},
        "JNJ":   {"price": 235, "mc": 565e9, "pe": 22, "beta": 0.5, "eps_t": 10.7, "eps_f": 11.5, "peg": 3.0, "w52h": 380, "w52l": 315},
        "AMD":   {"price": 165, "mc": 265e9, "pe": 28, "beta": 1.6, "eps_t": 4.90, "eps_f": 6.0, "peg": 1.2, "w52h": 500, "w52l": 120},
        # Additional stocks
        "TSLA":  {"price": 175, "mc": 560e9, "pe": 65, "beta": 2.0, "eps_t": 2.70, "eps_f": 4.0, "peg": 2.5, "w52h": 480, "w52l": 138},
        "NFLX":  {"price": 580, "mc": 250e9, "pe": 35, "beta": 1.3, "eps_t": 16.5, "eps_f": 20.0, "peg": 1.8, "w52h": 700, "w52l": 344},
        "CRM":   {"price": 270, "mc": 260e9, "pe": 45, "beta": 1.2, "eps_t": 6.0, "eps_f": 8.5, "peg": 2.5, "w52h": 370, "w52l": 195},
        "INTC":  {"price": 30, "mc": 125e9, "pe": 25, "beta": 1.0, "eps_t": 1.20, "eps_f": 1.80, "peg": 3.5, "w52h": 52, "w52l": 26},
        "BAC":   {"price": 37, "mc": 290e9, "pe": 12, "beta": 1.3, "eps_t": 3.10, "eps_f": 3.50, "peg": 1.5, "w52h": 48, "w52l": 30},
        "GS":    {"price": 420, "mc": 140e9, "pe": 14, "beta": 1.4, "eps_t": 30.0, "eps_f": 33.0, "peg": 1.8, "w52h": 550, "w52l": 290},
        "UNH":   {"price": 520, "mc": 480e9, "pe": 22, "beta": 0.6, "eps_t": 23.5, "eps_f": 26.0, "peg": 1.5, "w52h": 640, "w52l": 470},
        "WMT":   {"price": 185, "mc": 500e9, "pe": 28, "beta": 0.5, "eps_t": 6.6, "eps_f": 7.2, "peg": 2.5, "w52h": 220, "w52l": 155},
        "COST":  {"price": 720, "mc": 320e9, "pe": 45, "beta": 1.0, "eps_t": 16.0, "eps_f": 18.0, "peg": 2.8, "w52h": 920, "w52l": 580},
        "KO":    {"price": 62, "mc": 265e9, "pe": 22, "beta": 0.6, "eps_t": 2.82, "eps_f": 3.05, "peg": 3.2, "w52h": 74, "w52l": 52},
    }

    fb = FALLBACK.get(sym.upper().strip())
    if fb:
        return {**fb, "success": True}

    try:
        import ssl
        ctx = ssl.create_default_context()
        ctx.check_hostname = False
        ctx.verify_mode = ssl.CERT_NONE

        url = f"https://query1.finance.yahoo.com/v8/finance/chart/{sym}?interval=1d&range=5d"
        req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
        with urllib.request.urlopen(req, timeout=15, context=ctx) as r:
            d = json.loads(r.read().decode())
        meta = d["chart"]["result"][0]["meta"]

        info_url = f"https://query1.finance.yahoo.com/v10/finance/quoteSummary/{sym}?modules=defaultKeyStatistics"
        req2 = urllib.request.Request(info_url, headers={"User-Agent": "Mozilla/5.0"})
        with urllib.request.urlopen(req2, timeout=15, context=ctx) as r2:
            info = json.loads(r2.read().decode())

        stats = info["quoteSummary"]["result"][0].get("defaultKeyStatistics", {})

        return {
            "price": meta.get("regularMarketPrice", 0),
            "mc": meta.get("marketCap", 0),
            "pe": meta.get("trailingPE", 0),
            "beta": stats.get("beta", {}).get("raw", 1),
            "eps_t": stats.get("trailingEps", {}).get("raw", 0),
            "eps_f": stats.get("forwardEps", {}).get("raw", 0),
            "peg": stats.get("pegRatio", {}).get("raw", 0),
            "w52h": meta.get("fiftyTwoWeekHigh", 0),
            "w52l": meta.get("fiftyTwoWeekLow", 0),
            "success": True
        }
    except Exception as e:
        return {"success": False, "error": str(e)}

def calc_dcf_scenarios(q, fin):
    """DCF with Bull/Base/Bear scenarios"""
    eps_t = q["eps_t"]
    eps_f = q["eps_f"]
    beta = q["beta"]
    nm = fin["nm"]
    roic = fin["roic"]

    # Base case
    near_term_g = min((eps_f - eps_t) / eps_t, 0.30) if eps_t > 0 else 0.15

    scenarios = {
        "Bull": {"growth_mult": 1.3, "wacc_mult": 0.9},   # Higher growth, lower discount
        "Base": {"growth_mult": 1.0, "wacc_mult": 1.0},
        "Bear": {"growth_mult": 0.6, "wacc_mult": 1.2},   # Lower growth, higher discount
    }

    results = {}
    for name, params in scenarios.items():
        wacc = max(0.07, min(0.045 + 0.055 * beta * 0.5 * params["wacc_mult"], 0.15))
        g_term = min(roic * 0.4, 0.035)

        pv_fcf = 0
        for y in range(1, 6):
            g = near_term_g * params["growth_mult"] * (1 - y/5) + g_term * (y/5)
            g = min(max(g, 0), 0.25)
            eps_y = eps_t * (1 + g) ** y
            pv_fcf += eps_y * 0.7 / (1 + wacc) ** y

        eps_5y = eps_t * (1 + near_term_g * params["growth_mult"]) ** 5
        tv = eps_5y * 1.03 / (wacc - g_term)
        pv_tv = tv / (1 + wacc) ** 5

        results[name] = pv_fcf + pv_tv

    return results

def calc_5y_eps(q, fin):
    """Calculate 5-year EPS projection"""
    eps_t = q["eps_t"]
    eps_f = q["eps_f"]
    roic = fin["roic"]

    near_term_g = min((eps_f - eps_t) / eps_t, 0.30) if eps_t > 0 else 0.15
    g_term = min(roic * 0.4, 0.035)

    projections = []
    for y in range(1, 6):
        g = near_term_g * (1 - y/5) + g_term * (y/5)
        g = min(max(g, 0), 0.25)
        eps_y = eps_t * (1 + g) ** y
        projections.append({"year": y, "eps": eps_y, "growth": g * 100})

    return projections

def calc_risk_metrics(q):
    """Calculate risk metrics"""
    price = q["price"]
    beta = q["beta"]
    w52h = q["w52h"]
    w52l = q["w52l"]

    # VaR (Value at Risk) - 95% confidence
    daily_var = price * beta * 0.02  # 2% daily vol estimate
    var_95 = daily_var * 1.65  # 95% confidence

    # Max Drawdown from 52W high
    max_dd_pct = (price - w52h) / w52h * 100

    # Distance from 52W low (upside if buying now)
    from_low_pct = (price - w52l) / w52l * 100

    # Sharpe-like ratio (simplified)
    sharpe = (q["eps_t"] / price) / (beta * 0.20) if beta > 0 else 0

    # Position in 52W range
    range_pos = (price - w52l) / (w52h - w52l) * 100 if w52h > w52l else 50

    return {
        "var_95": var_95,
        "var_95_pct": var_95 / price * 100,
        "max_dd_pct": max_dd_pct,
        "from_low_pct": from_low_pct,
        "sharpe": sharpe,
        "range_pos": range_pos,
    }

def calc_technical(q):
    """Simple technical analysis"""
    price = q["price"]
    w52h = q["w52h"]
    w52l = q["w52l"]

    # Distance from key levels
    dist_from_high = (w52h - price) / price * 100
    dist_from_low = (price - w52l) / price * 100

    # Simple MA approximation (using 52W range)
    # Assume price oscillates around midpoint
    midpoint = (w52h + w52l) / 2
    ma_status = "Above Average" if price > midpoint else "Below Average"

    # Trend assessment
    range_pos = (price - w52l) / (w52h - w52l) if w52h > w52l else 0.5
    if range_pos > 0.7:
        trend = "Strong (Near High)"
    elif range_pos > 0.4:
        trend = "Neutral"
    else:
        trend = "Weak (Near Low)"

    return {
        "dist_from_high": dist_from_high,
        "dist_from_low": dist_from_low,
        "midpoint": midpoint,
        "ma_status": ma_status,
        "trend": trend,
        "range_pos": range_pos * 100,
        "support": w52l * 1.1,  # Approximate support at 52W low + 10%
        "resistance": w52h * 0.95,  # Approximate resistance at 52W high - 5%
    }

def calc_support_resistance(q):
    """Calculate support and resistance levels"""
    price = q["price"]
    w52h = q["w52h"]
    w52l = q["w52l"]

    range_size = w52h - w52l

    # Pivot points
    pivot = (w52h + w52l + price) / 3
    r1 = 2 * pivot - w52l
    r2 = pivot + range_size * 0.382
    r3 = pivot + range_size * 0.618
    r4 = w52h  # 52W high as final resistance

    s1 = 2 * pivot - w52h
    s2 = pivot - range_size * 0.382
    s3 = pivot - range_size * 0.618
    s4 = w52l  # 52W low as final support

    return {
        "pivot": pivot,
        "r1": r1, "r2": r2, "r3": r3, "r4": r4,
        "s1": s1, "s2": s2, "s3": s3, "s4": s4,
    }

def calc_earnings_calendar(sym):
    """Earnings calendar by sector"""
    CALENDAR = {
        "META": {"month": "Feb/May/Aug/Nov", "days_before": 14, "impact": "HIGH"},
        "NVDA": {"month": "Feb/May/Aug/Nov", "days_before": 21, "impact": "VERY HIGH"},
        "AAPL": {"month": "Jan/Apr/Jul/Oct", "days_before": 7, "impact": "MEDIUM"},
        "MSFT": {"month": "Jan/Apr/Jul/Oct", "days_before": 14, "impact": "MEDIUM"},
        "GOOGL": {"month": "Jan/Apr/Jul/Oct", "days_before": 14, "impact": "MEDIUM"},
        "AMZN": {"month": "Jan/Apr/Jul/Oct", "days_before": 21, "impact": "HIGH"},
        "JPM": {"month": "Jan/Apr/Jul/Oct", "days_before": 7, "impact": "LOW"},
        "JNJ": {"month": "Jan/Apr/Jul/Oct", "days_before": 7, "impact": "LOW"},
        "AMD": {"month": "Feb/May/Aug/Nov", "days_before": 21, "impact": "HIGH"},
        "TSLA": {"month": "Jan/Apr/Jul/Oct", "days_before": 28, "impact": "VERY HIGH"},
        "NFLX": {"month": "Jan/Apr/Jul/Oct", "days_before": 14, "impact": "HIGH"},
        "COST": {"month": "Sep/Dec/Mar/Jun", "days_before": 7, "impact": "LOW"},
        "KO": {"month": "Feb/May/Aug/Nov", "days_before": 7, "impact": "LOW"},
        "WMT": {"month": "Mar/Jun/Aug/Nov", "days_before": 7, "impact": "LOW"},
    }
    return CALENDAR.get(sym, {"month": "Quarterly", "days_before": 14, "impact": "MEDIUM"})

def calc_insider_sentiment(sym):
    """Mock insider trading sentiment"""
    INSIDER = {
        "META": {"buy_30d": 12, "sell_30d": 2, "sentiment": "Very Bullish"},
        "NVDA": {"buy_30d": 8, "sell_30d": 1, "sentiment": "Very Bullish"},
        "AAPL": {"buy_30d": 3, "sell_30d": 5, "sentiment": "Neutral"},
        "MSFT": {"buy_30d": 5, "sell_30d": 2, "sentiment": "Bullish"},
        "GOOGL": {"buy_30d": 4, "sell_30d": 1, "sentiment": "Bullish"},
        "AMZN": {"buy_30d": 6, "sell_30d": 3, "sentiment": "Bullish"},
        "JPM": {"buy_30d": 2, "sell_30d": 1, "sentiment": "Bullish"},
        "TSLA": {"buy_30d": 1, "sell_30d": 8, "sentiment": "Bearish"},
        "AMD": {"buy_30d": 5, "sell_30d": 2, "sentiment": "Bullish"},
    }
    return INSIDER.get(sym, {"buy_30d": 3, "sell_30d": 3, "sentiment": "Neutral"})

def calc_technical_indicators(q, fin):
    """Calculate RSI and other technical indicators"""
    price = q["price"]
    w52h = q["w52h"]
    w52l = q["w52l"]
    beta = q["beta"]

    # RSI estimate based on price position
    range_pos = (price - w52l) / (w52h - w52l) if w52h > w52l else 50
    rsi = 50 + (range_pos - 50) * 0.8  # Estimate RSI

    # MACD signal
    macd_signal = "Buy" if range_pos > 40 and range_pos < 70 else "Sell" if range_pos < 30 else "Neutral"

    # Moving averages (estimate based on typical trends)
    if range_pos > 60:
        ma_50 = price * 0.97
        ma_200 = price * 0.88
    elif range_pos < 40:
        ma_50 = price * 1.03
        ma_200 = price * 1.12
    else:
        ma_50 = price * 1.00
        ma_200 = price * 1.00

    price_vs_ma50 = "Above" if price > ma_50 else "Below"
    price_vs_ma200 = "Above" if price > ma_200 else "Below"

    # Bollinger Bands (estimate)
    range_size = w52h - w52l
    bb_upper = price + 0.1 * range_size
    bb_lower = price - 0.1 * range_size
    bb_width = (bb_upper - bb_lower) / price * 100

    # ATR (Average True Range) estimate
    atr = price * 0.02 * beta  # Scaled by volatility

    # Price Momentum (12-month)
    mom_12m = (price - w52l) / w52l * 100 - 50  # Relative position

    return {
        "rsi": min(max(rsi, 10), 90),
        "rsi_signal": "Overbought" if rsi > 70 else "Oversold" if rsi < 30 else "Neutral",
        "macd": macd_signal,
        "ma_50": ma_50,
        "ma_200": ma_200,
        "price_vs_ma50": price_vs_ma50,
        "price_vs_ma200": price_vs_ma200,
        "bb_upper": bb_upper,
        "bb_middle": price,
        "bb_lower": bb_lower,
        "bb_width": bb_width,
        "atr": atr,
        "mom_12m": mom_12m,
    }

def calc_balance_sheet_health(fin, q):
    """Balance sheet health analysis"""
    de = fin["de"]
    roe = fin["roe"]
    nm = fin["nm"]

    # Altman Z-Score approximation for public companies
    # Z = 1.2*X1 + 1.4*X2 + 3.3*X3 + 0.6*X4 + 1.0*X5
    # Simplified: X1=WC/TA, X2=RE/TA, X3=EBIT/TA, X4=Equity/DEBT, X5=Sales/TA
    z_score = 2.5 + 0.5 * (1 - de) + 0.3 * roe  # Simplified

    if z_score > 3:
        zd_status = "Safe Zone"
        zd_color = "Green"
    elif z_score > 1.8:
        zd_status = "Grey Zone"
        zd_color = "Yellow"
    else:
        zd_status = "Distress Zone"
        zd_color = "Red"

    # Current ratio estimate
    if de < 0.5:
        current_ratio = 2.0
    elif de < 1.5:
        current_ratio = 1.5
    else:
        current_ratio = 1.0

    # Debt capacity
    if de < 0.5:
        debt_capacity = "Strong"
        headroom = 100
    elif de < 1.5:
        debt_capacity = "Moderate"
        headroom = 50
    else:
        debt_capacity = "Limited"
        headroom = 20

    return {
        "z_score": z_score,
        "z_status": zd_status,
        "z_color": zd_color,
        "current_ratio": current_ratio,
        "debt_capacity": debt_capacity,
        "headroom_pct": headroom,
    }

def calc_competitors(sym):
    """Get competitor comparison"""
    COMPETITORS = {
        "META": ["GOOGL", "SNAP", "PINS"],
        "NVDA": ["AMD", "INTC", "QCOM"],
        "AAPL": ["SAMSUNG", "GOOGL", "MSFT"],
        "MSFT": ["AAPL", "GOOGL", "AMZN"],
        "GOOGL": ["META", "MSFT", "AMZN"],
        "AMZN": ["WMT", "COST", "TGT"],
        "JPM": ["BAC", "GS", "C", "WFC"],
        "JNJ": ["PFE", "UNH", "MRK"],
        "AMD": ["INTC", "NVDA", "QCOM"],
        "TSLA": ["F", "GM", "RIVN"],
        "NFLX": ["DIS", "WBD", "PARA"],
    }
    return COMPETITORS.get(sym, [])

def calc_market_cap_class(mc):
    """Classify market cap"""
    if mc >= 1e12:
        return "Mega Cap", ">$1T"
    elif mc >= 200e9:
        return "Large Cap", "$200B-$1T"
    elif mc >= 10e9:
        return "Mid Cap", "$10B-$200B"
    elif mc >= 2e9:
        return "Small Cap", "$2B-$10B"
    else:
        return "Micro Cap", "<$2B"

def calc_price_chart(q):
    """Generate ASCII price chart"""
    price = q["price"]
    w52h = q["w52h"]
    w52l = q["w52l"]

    # Create 20-bar chart
    bars = 20
    levels = []
    for i in range(bars):
        level_price = w52l + (w52h - w52l) * (bars - 1 - i) / (bars - 1)
        if price >= level_price:
            levels.append("█")
        else:
            levels.append("░")

    chart = "".join(levels)

    return {
        "chart": chart,
        "position": (bars - 1) * (price - w52l) / (w52h - w52l) if w52h > w52l else bars // 2,
        "w52l": w52l,
        "w52h": w52h,
    }

def calc_sector_comparison(sym, q, fin):
    """Compare stock metrics against sector average"""
    SECTOR_AVG = {
        "Tech": {"pe": 25, "peg": 1.5, "roe": 0.20, "de": 0.8, "fcf_yield": 0.025},
        "Finance": {"pe": 12, "peg": 1.2, "roe": 0.15, "de": 2.0, "fcf_yield": 0.045},
        "Healthcare": {"pe": 18, "peg": 2.0, "roe": 0.18, "de": 0.5, "fcf_yield": 0.035},
        "Consumer": {"pe": 20, "peg": 2.5, "roe": 0.25, "de": 0.7, "fcf_yield": 0.040},
    }

    TECH = ["META", "NVDA", "AMD", "INTC", "QCOM", "CRM", "ORCL", "MSFT", "GOOGL", "AMZN"]
    FINANCE = ["JPM", "BAC", "GS", "UNH"]
    HEALTHCARE = ["JNJ", "UNH"]
    CONSUMER = ["WMT", "COST", "KO", "PG"]

    if sym in TECH:
        sector = "Tech"
    elif sym in FINANCE:
        sector = "Finance"
    elif sym in HEALTHCARE:
        sector = "Healthcare"
    else:
        sector = "Consumer"

    avg = SECTOR_AVG[sector]

    pe_vs = "Cheaper" if q["pe"] < avg["pe"] * 0.9 else "Expensive" if q["pe"] > avg["pe"] * 1.1 else "Fair"
    peg_vs = "Cheaper" if q["peg"] < avg["peg"] * 0.9 else "Expensive" if q["peg"] > avg["peg"] * 1.1 else "Fair"
    roe_vs = "Better" if fin["roe"] > avg["roe"] else "Below Avg"
    fcf_vs = "Better" if fin.get("fcf_yield", 0) > avg["fcf_yield"] else "Below Avg"

    return {
        "sector": sector,
        "avg_pe": avg["pe"],
        "stock_pe": q["pe"],
        "pe_vs": pe_vs,
        "avg_peg": avg["peg"],
        "stock_peg": q["peg"],
        "peg_vs": peg_vs,
        "roe_vs": roe_vs,
        "fcf_vs": fcf_vs,
    }

def calc_dividend_analysis(q, fin, cf):
    """Dividend health analysis"""
    div_yield = cf.get("div_yield", 0)
    payout = fin.get("payout", 0)

    if div_yield < 0.01:
        status = "No Dividend"
        health = "N/A"
    elif payout < 0.40:
        status = "Sustainable"
        health = "Excellent"
    elif payout < 0.60:
        status = "Sustainable"
        health = "Good"
    elif payout < 0.80:
        status = "Moderate"
        health = "Fair"
    else:
        status = "At Risk"
        health = "Poor"

    return {
        "div_yield": div_yield,
        "payout": payout,
        "status": status,
        "health": health,
    }

def calc_price_target_probability(price, target, fv_low, fv_high, dcf_base):
    """Calculate probability-weighted price target"""
    # Conservative (20%), Base (50%), Optimistic (30%)
    conservative = fv_low
    base = target
    optimistic = fv_high

    expected = conservative * 0.20 + base * 0.50 + optimistic * 0.30

    # Upside/downside from current
    upside_pct = (expected - price) / price * 100

    return {
        "expected": expected,
        "upside_pct": upside_pct,
        "conservative": conservative,
        "base": base,
        "optimistic": optimistic,
    }



def calc_analyst_consensus(sym):
    """Get analyst consensus details"""
    return ANALYST_RATINGS.get(sym, {"buy": 50, "hold": 35, "sell": 15})

def calc_institutional_ownership(sym):
    """Institutional ownership estimates by sector"""
    INST_OWNERSHIP = {
        "META": 78, "NVDA": 75, "AAPL": 60, "MSFT": 65, "GOOGL": 65,
        "AMZN": 55, "JPM": 70, "JNJ": 55, "AMD": 65, "TSLA": 45,
        "NFLX": 70, "INTC": 60, "BAC": 65, "KO": 60, "WMT": 55,
    }
    return INST_OWNERSHIP.get(sym, 60)

def calc_quality_score(q, fin):
    """Calculate quality score (quality = consistency of profits)"""
    score = 50

    # ROE quality
    if fin["roic"] > fin["roe"]:
        score += 10  # ROIC > ROE means efficient capital allocation
    else:
        score -= 5

    # Margin stability
    if fin["nm"] > 0.15:
        score += 10
    elif fin["nm"] > 0.08:
        score += 5

    # FCF quality
    if fin.get("fcf_yield", 0) > 0.03:
        score += 10
    elif fin.get("fcf_yield", 0) > 0:
        score += 5

    # Debt management
    if fin["de"] < 0.5:
        score += 10
    elif fin["de"] < 1.0:
        score += 5
    elif fin["de"] > 2.0:
        score -= 10

    # Profitability consistency
    if fin["roe"] > 0.20:
        score += 10
    elif fin["roe"] > 0.15:
        score += 5

    return min(max(score, 0), 100)

def calc_value_score(q, fin, val_pct):
    """Calculate value score"""
    score = 50

    # P/E vs historical
    if val_pct["percentile"] < 30:
        score += 20
    elif val_pct["percentile"] < 50:
        score += 10
    elif val_pct["percentile"] > 70:
        score -= 15

    # PEG ratio
    if q["peg"] < 1.0:
        score += 15
    elif q["peg"] < 1.5:
        score += 8
    elif q["peg"] > 2.5:
        score -= 10

    # P/FCF
    if fin.get("fcf_yield", 0) > 0.06:
        score += 15
    elif fin.get("fcf_yield", 0) > 0.04:
        score += 8
    elif fin.get("fcf_yield", 0) < 0.02:
        score -= 5

    return min(max(score, 0), 100)

def calc_momentum_score(q, mom):
    """Calculate momentum score"""
    score = 50

    # Range position
    if mom["range_pos"] > 80:
        score += 10
    elif mom["range_pos"] < 20:
        score += 5
    elif mom["range_pos"] > 90 or mom["range_pos"] < 10:
        score -= 10  # Extreme positions can reverse

    # Sentiment
    if mom["momentum"] in ["Extreme Fear"]:
        score += 15
    elif mom["momentum"] in ["Fear"]:
        score += 8
    elif mom["momentum"] in ["Greed", "Extreme Greed"]:
        score -= 10

    return min(max(score, 0), 100)

def calc_growth_score(q, fin):
    """Calculate growth score"""
    score = 50

    # Revenue growth
    rev_g = fin.get("rev_g", 0)
    if rev_g > 0.25:
        score += 20
    elif rev_g > 0.15:
        score += 15
    elif rev_g > 0.10:
        score += 10
    elif rev_g > 0.05:
        score += 5
    elif rev_g < 0:
        score -= 10

    # EPS growth
    if q["eps_f"] > q["eps_t"] * 1.20:
        score += 15
    elif q["eps_f"] > q["eps_t"] * 1.10:
        score += 10
    elif q["eps_f"] < q["eps_t"]:
        score -= 10

    # PEG (growth at reasonable price)
    if q["peg"] < 1.0:
        score += 15
    elif q["peg"] < 1.5:
        score += 8

    return min(max(score, 0), 100)

def calc_technical(q):
    """Calculate basic technical metrics"""
    w52h = q.get("w52h", q["price"] * 1.3)
    w52l = q.get("w52l", q["price"] * 0.7)
    price = q["price"]

    dist_from_high = (w52h - price) / w52h * 100
    dist_from_low = (price - w52l) / w52l * 100
    midpoint = (w52h + w52l) / 2

    ma_status = "Above Average" if price > midpoint else "Below Average"
    range_pos = (price - w52l) / (w52h - w52l) if w52h > w52l else 0.5
    if range_pos > 0.7:
        trend = "Strong (Near High)"
    elif range_pos > 0.4:
        trend = "Neutral"
    else:
        trend = "Weak (Near Low)"

    return {
        "dist_from_high": dist_from_high,
        "dist_from_low": dist_from_low,
        "midpoint": midpoint,
        "ma_status": ma_status,
        "trend": trend,
        "range_pos": range_pos * 100,
        "support": w52l * 1.1,
        "resistance": w52h * 0.95,
    }

def calc_valuation_percentile(q, fin):
    """Calculate where current valuation sits in historical context"""
    pe_hist = PE_HIST.get(sym, {"curr": q["pe"], "low": 15, "high": 35, "avg": 25})

    curr_pe = q["pe"]
    low_pe = pe_hist["low"]
    high_pe = pe_hist["high"]
    avg_pe = pe_hist["avg"]

    # Percentile position
    if high_pe > low_pe:
        pct = (curr_pe - low_pe) / (high_pe - low_pe) * 100
        pct = max(0, min(100, pct))
    else:
        pct = 50

    # Assessment
    if pct < 20:
        assessment = "Deep Value (Historical Discount)"
    elif pct < 40:
        assessment = "Below Average (Attractive)"
    elif pct < 60:
        assessment = "Fair Value"
    elif pct < 80:
        assessment = "Above Average (Expensive)"
    else:
        assessment = "Premium (Very Expensive)"

    return {
        "curr_pe": curr_pe,
        "low_pe": low_pe,
        "avg_pe": avg_pe,
        "high_pe": high_pe,
        "percentile": pct,
        "assessment": assessment,
    }

def calc_cash_flow_analysis(q, fin):
    """Analyze cash flow quality"""
    price = q["price"]
    fcf_yield = fin.get("fcf_yield", 0.03)
    div_yield = fin.get("div_yield", 0)
    payout = fin.get("payout", 0)

    # FCF yield vs bond yield
    bond_yield = 0.045
    fcf_spread = fcf_yield - bond_yield

    # FCF收益率
    fcf_value = price * fcf_yield

    # 股息安全度
    if payout > 0 and div_yield > 0:
        implied_payout = div_yield / (fcf_yield + 0.001) * 100
        div_safe = "Safe" if implied_payout < 50 else "Moderate" if implied_payout < 80 else "Risky"
    else:
        implied_payout = 0
        div_safe = "N/A"

    return {
        "fcf_yield": fcf_yield,
        "fcf_value": fcf_value,
        "div_yield": div_yield,
        "payout": payout,
        "fcf_spread": fcf_spread,
        "div_safe": div_safe,
        "implied_payout": implied_payout,
    }

def calc_momentum(q):
    """Calculate price momentum indicators"""
    price = q["price"]
    w52h = q["w52h"]
    w52l = q["w52l"]

    # YTD return approximation
    ytd_approx = (price - w52l) / w52l * 100 if w52l > 0 else 0

    # Momentum score (0-100)
    range_pos = (price - w52l) / (w52h - w52l) if w52h > w52l else 0.5

    if range_pos > 0.9:
        mom = "Extreme Greed"
    elif range_pos > 0.75:
        mom = "Greed"
    elif range_pos > 0.55:
        mom = "Neutral"
    elif range_pos > 0.4:
        mom = "Fear"
    else:
        mom = "Extreme Fear"

    return {
        "range_pos": range_pos * 100,
        "ytd_approx": ytd_approx,
        "momentum": mom,
    }

def calc_value_trap_detection(q, fin):
    """Detect value traps (cheap stocks that stay cheap)"""
    warnings = []

    # Low P/E trap
    if q["pe"] < 15 and fin["roe"] < 0.15:
        warnings.append({"type": "Low P/E, Low ROE", "risk": "Value Trap", "detail": "Cheap valuation may reflect deteriorating fundamentals"})

    # Declining ROE
    if fin["roe"] < fin["roic"]:
        warnings.append({"type": "ROE < ROIC", "risk": "Leverage Dependent", "detail": "Returns driven by debt, not operations"})

    # Negative FCF
    if fin.get("fcf_yield", 0) < 0:
        warnings.append({"type": "Negative FCF", "risk": "Cash Burn", "detail": "Not generating free cash flow"})

    # Declining margins
    if fin["nm"] < 0.10 and fin["nm"] < 0.15:
        warnings.append({"type": "Thin Margins", "risk": "Competition", "detail": "Low profitability vulnerable to competition"})

    return warnings if warnings else None

def calc_fair_value_range(dcf_scenarios, analyst_target):
    """Calculate fair value range from scenarios"""
    values = [dcf_scenarios["Bear"], analyst_target, dcf_scenarios["Bull"]]
    return min(values), max(values)

def score_stock(q, fin):
    """Improved scoring with more dimensions"""
    s = 50
    # Profitability (max 35)
    s += 10 if fin["gm"] > 0.60 else 5 if fin["gm"] > 0.40 else 0
    s += 15 if fin["nm"] > 0.25 else 10 if fin["nm"] > 0.15 else 0
    s += 10 if fin["roe"] > 0.30 else 5 if fin["roe"] > 0.20 else 0
    # Efficiency (max 15)
    s += 10 if fin["roic"] > 0.25 else 5 if fin["roic"] > 0.15 else 0
    s += 5 if fin["roic"] > fin["roe"] * 0.8 else 0  # Good capital allocation
    # Balance Sheet (max 15)
    s += 10 if fin["de"] < 0.8 else 5 if fin["de"] < 1.5 else 0
    # Risk Adjustments (max -30)
    s -= 15 if q["beta"] > 1.5 else 5 if q["beta"] > 1.2 else 0
    s -= 10 if q["peg"] > 2.5 else 0
    # Valuation Bonus (max +15)
    if q["pe"] < 20: s += 10
    elif q["pe"] < 25: s += 5
    if q["peg"] < 1.0: s += 5
    return min(max(s, 0), 100)

def rating(upside):
    if upside > 30: return "STRONG_BUY"
    if upside > 15: return "BUY"
    if upside > 0: return "HOLD"
    return "SELL"

def risk_level(q, fin):
    risk = q["beta"] * 20 + fin["de"] * 10
    if risk < 30: return "LOW"
    if risk < 50: return "MEDIUM"
    return "HIGH"

def generate_report(sym):
    sym = sym.upper().strip()
    print(f"\n{'='*60}\nStock PRO v10.0 - Analyzing {sym}\n{'='*60}")

    q = fetch(sym)
    if not q.get("success"):
        print(f"Failed: {q.get('error', 'Unknown')}")
        return None

    fin = FD.get(sym, FD["AAPL"])
    analyst = ANALYST.get(sym, {"target": q["price"], "rating": "N/A"})

    price = q["price"]
    analyst_target = analyst["target"]

    # Get DCF scenarios
    dcf_scenarios = calc_dcf_scenarios(q, fin)
    dcf_base = dcf_scenarios["Base"]

    # Weighted target: Analyst 60%, DCF Base 40%
    target = analyst_target * 0.6 + dcf_base * 0.4
    upside = (target - price) / price * 100

    sc = score_stock(q, fin)
    rating_str = rating(upside)
    risk = risk_level(q, fin)
    pe_f = price / q["eps_f"] if q["eps_f"] > 0 else 0

    # Additional metrics
    eps_proj = calc_5y_eps(q, fin)
    risk_m = calc_risk_metrics(q)
    tech = calc_technical(q)
    val_pct = calc_valuation_percentile(q, fin)
    cf = calc_cash_flow_analysis(q, fin)
    mom = calc_momentum(q)
    traps = calc_value_trap_detection(q, fin)
    fv_low, fv_high = calc_fair_value_range(dcf_scenarios, analyst_target)
    sr_levels = calc_support_resistance(q)
    analyst_ratings = calc_analyst_consensus(sym)
    inst_own = calc_institutional_ownership(sym)

    # New scoring system
    quality = calc_quality_score(q, fin)
    value = calc_value_score(q, fin, val_pct)
    momentum_s = calc_momentum_score(q, mom)
    growth = calc_growth_score(q, fin)

    # Combined score
    sc = int(quality * 0.30 + value * 0.25 + momentum_s * 0.20 + growth * 0.25)

    # New analysis modules
    earnings_cal = calc_earnings_calendar(sym)
    insider = calc_insider_sentiment(sym)
    tech_ind = calc_technical_indicators(q, fin)
    sector = calc_sector_comparison(sym, q, fin)
    div_info = calc_dividend_analysis(q, fin, cf)
    prob_target = calc_price_target_probability(price, target, fv_low, fv_high, dcf_base)
    balance = calc_balance_sheet_health(fin, q)
    competitors = calc_competitors(sym)
    cap_class, cap_range = calc_market_cap_class(q["mc"])
    price_chart = calc_price_chart(q)

    print(f"\nPrice: ${price:.2f}")
    print(f"EPS: Trailing ${q['eps_t']:.2f} | Forward ${q['eps_f']:.2f}")
    print(f"P/E: Current {q['pe']:.1f}x | Forward {pe_f:.1f}x")
    print(f"PEG: {q['peg']:.2f} | Beta: {q['beta']:.2f}")
    print(f"\nDCF Scenarios: Bull=${dcf_scenarios['Bull']:.0f} | Base=${dcf_base:.0f} | Bear=${dcf_scenarios['Bear']:.0f}")
    print(f"Analyst Target: ${analyst_target:.0f} ({analyst['rating']})")
    print(f"Combined Target: ${target:.2f} ({upside:+.1f}%)")
    print(f"\nRating: {rating_str} | Score: {sc}/100 | Risk: {risk}")
    print(f"Scores: Quality={quality} | Value={value} | Momentum={momentum_s} | Growth={growth}")
    print(f"5Y EPS: ${eps_proj[-1]['eps']:.2f} ({eps_proj[-1]['growth']:.1f}% CAGR)")
    print(f"P/E Percentile: {val_pct['percentile']:.0f}% ({val_pct['assessment']})")
    print(f"FCF Yield: {cf['fcf_yield']*100:.1f}% | Div Yield: {cf['div_yield']*100:.2f}%")
    print(f"Institutional Ownership: {inst_own}%")

    date = datetime.now().strftime("%Y-%m-%d")
    md_path = OUTPUT / f"{sym}_{date}.md"

    with open(md_path, "w", encoding="utf-8") as f:
        f.write(f"# {sym} Stock Analysis Report\n\n")
        f.write(f"**Generated:** {date} | **Version:** v9.0\n\n")
        f.write("---\n\n")

        # Rating Badge
        badge = {"STRONG_BUY": "🟢", "BUY": "🔵", "HOLD": "🟡", "SELL": "🔴"}.get(rating_str, "⚪")
        f.write(f"## {badge} Investment Rating: {rating_str}\n\n")
        f.write(f"| Key Metrics | Value |\n")
        f.write(f"|-------------|-------|\n")
        f.write(f"| Current Price | **${price:.2f}** |\n")
        f.write(f"| 12-Month Target | **${target:.2f}** |\n")
        f.write(f"| Upside Potential | **{upside:+.1f}%** |\n")
        f.write(f"| Fundamental Score | **{sc}/100** |\n")
        f.write(f"| Risk Level | {risk} |\n\n")
        f.write("---\n\n")

        # Executive Summary Table
        f.write("## 1. Executive Summary\n\n")
        f.write(f"| Metric | Value | Assessment |\n")
        f.write(f"|--------|-------|------------|\n")
        f.write(f"| Current Price | ${price:.2f} | - |\n")
        f.write(f"| Analyst Target | ${analyst_target:.0f} | {analyst['rating']} |\n")
        f.write(f"| DCF Base Case | ${dcf_base:.2f} | Intrinsic Value |\n")
        f.write(f"| Combined Target | ${target:.2f} | **Best Estimate** |\n")
        f.write(f"| Upside/Downside | {upside:+.1f}% | {rating_str} |\n")
        f.write(f"| Score | {sc}/100 | - |\n")
        f.write(f"| Risk | {risk} | - |\n\n")

        # Score Breakdown
        f.write("### 1.1 Score Breakdown\n\n")
        f.write(f"| Dimension | Score | Weight | Contribution |\n")
        f.write(f"|-----------|-------|--------|--------------|\n")
        f.write(f"| Quality | {quality} | 30% | {quality*0.30:.0f} |\n")
        f.write(f"| Value | {value} | 25% | {value*0.25:.0f} |\n")
        f.write(f"| Growth | {growth} | 25% | {growth*0.25:.0f} |\n")
        f.write(f"| Momentum | {momentum_s} | 20% | {momentum_s*0.20:.0f} |\n")
        f.write(f"| **Combined** | **{sc}** | 100% | - |\n\n")

        # Score gauge
        score_bar = "█" * int(sc/5) + "░" * (20 - int(sc/5))
        f.write(f"Overall Score: [{score_bar}] {sc}/100\n\n")

        f.write("---\n\n")
        f.write("## 2. Valuation Analysis\n\n")
        f.write("### 2.1 Valuation Multiples\n\n")
        f.write(f"| Metric | Value | Assessment |\n")
        f.write(f"|--------|-------|------------|\n")
        pe_assess = "Expensive" if q["pe"] > 30 else "Fair" if q["pe"] > 15 else "Cheap"
        f.write(f"| Trailing P/E | {q['pe']:.1f}x | {pe_assess} |\n")
        pe_f_assess = "Expensive" if pe_f > 25 else "Fair" if pe_f > 15 else "Cheap"
        f.write(f"| Forward P/E | {pe_f:.1f}x | {pe_f_assess} |\n")
        peg_assess = "Overvalued" if q["peg"] > 2 else "Fair" if q["peg"] > 1 else "Undervalued"
        f.write(f"| PEG Ratio | {q['peg']:.2f} | {peg_assess} |\n\n")
        f.write("### 2.2 DCF Scenario Analysis\n\n")
        f.write(f"| Scenario | DCF Value | vs Current |\n")
        f.write(f"|----------|-----------|-----------|\n")
        for name, val in dcf_scenarios.items():
            diff = (val - price) / price * 100
            color = "🟢" if diff > 0 else "🔴"
            f.write(f"| {name} | ${val:.2f} | {color} {diff:+.1f}% |\n")
        f.write(f"| **Analyst** | **${analyst_target:.0f}** | 🟢 {(analyst_target-price)/price*100:+.1f}% |\n\n")

        f.write("### 2.3 Target Price Derivation\n\n")
        f.write(f"| Method | Target Price | Weight |\n")
        f.write(f"|--------|-------------|--------|\n")
        f.write(f"| Analyst Consensus | ${analyst_target:.0f} | 60% |\n")
        f.write(f"| DCF Base Case | ${dcf_base:.2f} | 40% |\n")
        f.write(f"| **Combined** | **${target:.2f}** | 100% |\n\n")
        f.write("---\n\n")

        # 5-Year EPS Projection
        f.write("## 3. Earnings Outlook (5-Year)\n\n")
        f.write("| Year | EPS | Growth | Notes |\n")
        f.write("|------|-----|--------|-------|\n")
        for proj in eps_proj:
            note = "High Growth" if proj["growth"] > 20 else "Stable" if proj["growth"] > 10 else "Mature"
            f.write(f"| Y{proj['year']} | ${proj['eps']:.2f} | {proj['growth']:.1f}% | {note} |\n")
        f.write(f"| **CAGR** | - | **{eps_proj[-1]['growth']:.1f}%** | 5-Year |\n\n")
        f.write("---\n\n")

        # Risk Metrics
        f.write("## 4. Risk Analysis\n\n")
        f.write(f"| Metric | Value | Interpretation |\n")
        f.write(f"|--------|-------|---------------|\n")
        f.write(f"| VaR (95%) | ${risk_m['var_95']:.2f} | {risk_m['var_95_pct']:.1f}% of price |\n")
        f.write(f"| Max Drawdown Risk | {risk_m['max_dd_pct']:.1f}% | From 52W high |\n")
        f.write(f"| Sharpe Ratio (est) | {risk_m['sharpe']:.2f} | Risk-adjusted return |\n")
        f.write(f"| 52W Position | {risk_m['range_pos']:.0f}% | {'Near High' if risk_m['range_pos'] > 70 else 'Near Low' if risk_m['range_pos'] < 30 else 'Mid-range'} |\n\n")
        f.write(f"| 52W Range | ${q['w52l']:.0f} - ${q['w52h']:.0f} | Current: ${price:.2f} |\n\n")

        # Technical Analysis
        f.write("### 4.1 Technical Assessment\n\n")
        f.write(f"| Metric | Value |\n")
        f.write(f"|--------|-------|\n")
        f.write(f"| Trend | {tech['trend']} |\n")
        f.write(f"| Distance from 52W High | {tech['dist_from_high']:.1f}% |\n")
        f.write(f"| Distance from 52W Low | +{tech['dist_from_low']:.1f}% |\n")
        f.write(f"| Position vs Average | {tech['ma_status']} |\n")
        f.write(f"| Fair Value Midpoint | ${tech['midpoint']:.2f} |\n\n")

        # Support/Resistance
        f.write("### 4.2 Support & Resistance Levels\n\n")
        f.write(f"| Level | Price | Distance |\n")
        f.write(f"|-------|-------|----------|\n")
        f.write(f"| Resistance R4 | ${sr_levels['r4']:.2f} | +{(sr_levels['r4']-price)/price*100:.1f}% |\n")
        f.write(f"| Resistance R3 | ${sr_levels['r3']:.2f} | +{(sr_levels['r3']-price)/price*100:.1f}% |\n")
        f.write(f"| Resistance R2 | ${sr_levels['r2']:.2f} | +{(sr_levels['r2']-price)/price*100:.1f}% |\n")
        f.write(f"| Resistance R1 | ${sr_levels['r1']:.2f} | +{(sr_levels['r1']-price)/price*100:.1f}% |\n")
        f.write(f"| **Current Price** | **${price:.2f}** | 0.0% |\n")
        f.write(f"| Support S1 | ${sr_levels['s1']:.2f} | {(sr_levels['s1']-price)/price*100:.1f}% |\n")
        f.write(f"| Support S2 | ${sr_levels['s2']:.2f} | {(sr_levels['s2']-price)/price*100:.1f}% |\n")
        f.write(f"| Support S3 | ${sr_levels['s3']:.2f} | {(sr_levels['s3']-price)/price*100:.1f}% |\n\n")

        f.write("---\n\n")
        f.write("## 5. Financial Health\n\n")
        f.write(f"| Metric | Value | Rating |\n")
        f.write(f"|--------|-------|--------|\n")
        gm_r = "Excellent" if fin["gm"] > 0.60 else "Good" if fin["gm"] > 0.40 else "Average"
        f.write(f"| Gross Margin | {fin['gm']*100:.1f}% | {gm_r} |\n")
        nm_r = "Excellent" if fin["nm"] > 0.25 else "Good" if fin["nm"] > 0.15 else "Average"
        f.write(f"| Net Margin | {fin['nm']*100:.1f}% | {nm_r} |\n")
        roe_r = "Excellent" if fin["roe"] > 0.30 else "Good" if fin["roe"] > 0.20 else "Average"
        f.write(f"| ROE | {fin['roe']*100:.1f}% | {roe_r} |\n")
        roic_r = "Excellent" if fin["roic"] > 0.25 else "Good" if fin["roic"] > 0.15 else "Average"
        f.write(f"| ROIC | {fin['roic']*100:.1f}% | {roic_r} |\n")
        de_r = "Low" if fin["de"] < 0.8 else "Medium" if fin["de"] < 1.5 else "High"
        f.write(f"| Debt/Equity | {fin['de']:.1f}x | {de_r} |\n")
        f.write(f"| Beta | {q['beta']:.2f} | Volatile |\n")
        f.write(f"| 52W Range | ${q['w52l']:.0f}-${q['w52h']:.0f} | - |\n\n")

        # Cash Flow Analysis
        f.write("### 5.1 Cash Flow Analysis\n\n")
        f.write(f"| Metric | Value | Assessment |\n")
        f.write(f"|--------|-------|------------|\n")
        f.write(f"| FCF Yield | {cf['fcf_yield']*100:.1f}% | {'Strong' if cf['fcf_yield'] > 0.04 else 'Average'} |\n")
        f.write(f"| FCF Value | ${cf['fcf_value']:.2f}/share | Implied annual FCF |\n")
        f.write(f"| FCF vs Bonds | {cf['fcf_spread']*100:+.1f}% | {'Attractive' if cf['fcf_spread'] > 0 else 'Below bond yield'} |\n")
        f.write(f"| Dividend Yield | {cf['div_yield']*100:.2f}% | {'Paid' if cf['div_yield'] > 0 else 'No dividend'} |\n")
        if cf['div_yield'] > 0:
            f.write(f"| Payout Ratio | {cf['payout']*100:.0f}% | {cf['div_safe']} |\n")
        f.write(f"| EV/EBITDA | {fin.get('ev_ebitda', 'N/A')}x | {'High' if fin.get('ev_ebitda', 0) > 25 else 'Moderate' if fin.get('ev_ebitda', 0) > 15 else 'Low'} |\n\n")

        # Balance Sheet Health
        f.write("### 5.2 Balance Sheet Health\n\n")
        f.write(f"| Metric | Value | Status |\n")
        f.write(f"|--------|-------|--------|\n")
        z_color_icon = {"Green": "✓", "Yellow": "⚠", "Red": "✗"}.get(balance['z_color'], "")
        f.write(f"| Altman Z-Score | {balance['z_score']:.1f} | {z_color_icon} {balance['z_status']} |\n")
        f.write(f"| Current Ratio | {balance['current_ratio']:.1f}x | {'Strong' if balance['current_ratio'] > 1.5 else 'Adequate'} |\n")
        f.write(f"| Debt Capacity | {balance['debt_capacity']} | Headroom: {balance['headroom_pct']}% |\n")
        f.write(f"| Market Cap | ${q['mc']/1e9:.0f}B | {cap_class} |\n\n")

        # 52W Price Chart
        f.write("### 5.3 Price Position Chart\n\n")
        f.write(f"```\n")
        f.write(f"52W High ${q['w52h']:.0f} ┬\n")
        chart_line = "                    " + "█" * int(price_chart['position']) + "●\n"
        f.write(f"{chart_line}")
        f.write(f"                    {price_chart['chart']}\n")
        f.write(f"52W Low  ${q['w52l']:.0f} ┴\n")
        f.write(f"```\n")
        f.write(f"Current: ${price:.2f} ({(price - q['w52l'])/(q['w52h'] - q['w52l'])*100:.0f}% of range)\n\n")

        # Historical Valuation
        f.write("---\n\n")
        f.write("## 6. Historical Valuation Analysis\n\n")
        f.write("| Metric | Value | Context |\n")
        f.write(f"|--------|-------|---------|\n")
        f.write(f"| Current P/E | {val_pct['curr_pe']:.0f}x | - |\n")
        f.write(f"| 5Y Low P/E | {val_pct['low_pe']:.0f}x | Historical bottom |\n")
        f.write(f"| 5Y Avg P/E | {val_pct['avg_pe']:.0f}x | Historical mean |\n")
        f.write(f"| 5Y High P/E | {val_pct['high_pe']:.0f}x | Historical peak |\n")
        f.write(f"| **Percentile** | **{val_pct['percentile']:.0f}%** | {val_pct['assessment']} |\n\n")

        # Visual P/E gauge
        bar_len = 30
        filled = int(val_pct['percentile'] / 100 * bar_len)
        bar = "█" * filled + "░" * (bar_len - filled)
        f.write(f"P/E Position: [{bar}]\n\n")
        f.write(f"```\n")
        f.write(f"  {val_pct['low_pe']:.0f}x (Low) ---- {val_pct['avg_pe']:.0f}x (Avg) ---- {val_pct['high_pe']:.0f}x (High)\n")
        f.write(f"                 ↑\n")
        f.write(f"            Current: {val_pct['curr_pe']:.0f}x ({val_pct['percentile']:.0f}%)\n")
        f.write(f"```\n\n")

        # Analyst Consensus
        f.write("### 6.1 Analyst Consensus\n\n")
        f.write(f"| Rating | % of Analysts | Description |\n")
        f.write(f"|--------|---------------|-------------|\n")
        f.write(f"| Buy/Outperform | {analyst_ratings['buy']}% | Strong conviction |\n")
        f.write(f"| Hold | {analyst_ratings['hold']}% | Neutral outlook |\n")
        f.write(f"| Sell/Underperform | {analyst_ratings['sell']}% | Negative view |\n\n")

        # Visual rating distribution
        buy_bar = "█" * int(analyst_ratings['buy'] / 5)
        hold_bar = "█" * int(analyst_ratings['hold'] / 5)
        sell_bar = "█" * int(analyst_ratings['sell'] / 5)
        f.write(f"Rating Distribution: [{buy_bar}]({analyst_ratings['buy']}% Buy) [{hold_bar}]({analyst_ratings['hold']}% Hold) [{sell_bar}]({analyst_ratings['sell']}% Sell)\n\n")

        f.write(f"| Metric | Value |\n")
        f.write(f"|--------|-------|\n")
        f.write(f"| Analysts Covering | {analyst.get('num', 'N/A')} | - |\n")
        f.write(f"| Consensus Rating | {analyst['rating']} | - |\n")
        f.write(f"| 12M Target | ${analyst_target:.0f} | {upside:+.1f}% upside |\n\n")

        # Institutional Ownership
        f.write("### 6.2 Institutional Ownership\n\n")
        f.write(f"| Metric | Value |\n")
        f.write(f"|--------|-------|\n")
        f.write(f"| Inst. Ownership | {inst_own}% | {'High (stable)' if inst_own > 65 else 'Medium' if inst_own > 50 else 'Low (volatile)'} |\n\n")

        # Momentum
        f.write("### 6.1 Market Sentiment & Momentum\n\n")
        f.write(f"| Indicator | Value | Interpretation |\n")
        f.write(f"|-----------|-------|----------------|\n")
        f.write(f"| Sentiment | {mom['momentum']} | Market positioning |\n")
        f.write(f"| 52W Position | {mom['range_pos']:.0f}% | {'Near top' if mom['range_pos'] > 75 else 'Near bottom' if mom['range_pos'] < 25 else 'Mid-range'} |\n\n")

        # Value Trap Warnings
        if traps:
            f.write("### 6.2 ⚠️ Value Trap Detection\n\n")
            f.write("| Warning | Risk Type | Details |\n")
            f.write("|---------|-----------|--------|\n")
            for trap in traps:
                f.write(f"| {trap['type']} | {trap['risk']} | {trap['detail']} |\n")
            f.write("\n")

        # Sector Comparison
        peers = {
            "META":  [("GOOGL", 25, 0.56), ("SNAP", 999, 0.44)],
            "NVDA":  [("AMD", 28, 0.47), ("INTC", 999, 0.40)],
            "AAPL":  [("MSFT", 36, 0.70), ("GOOGL", 25, 0.56)],
            "MSFT":  [("AAPL", 32, 0.47), ("GOOGL", 25, 0.56)],
            "JPM":   [("BAC", 11, 0.58), ("WFC", 10, 0.50)],
            "JNJ":   [("PFE", 15, 0.60), ("UNH", 22, 0.21)],
            "GOOGL": [("META", 35, 0.81), ("AMZN", 45, 0.48)],
            "AMZN":  [("WMT", 28, 0.25), ("TGT", 14, 0.28)],
        }

        if sym in peers:
            f.write("---\n\n")
            f.write("## 7. Sector Comparison\n\n")
            f.write("| Peer | P/E | Gross Margin | vs Subject |\n")
            f.write("|------|-----|--------------|----------|\n")
            my_pe = q["pe"]
            my_gm = fin["gm"]
            for peer, peer_pe, peer_gm in peers[sym]:
                pe_diff = "Cheaper" if peer_pe > my_pe else "Expensive"
                gm_diff = "Higher" if peer_gm > my_gm else "Lower"
                f.write(f"| {peer} | {peer_pe}x | {peer_gm*100:.0f}% | {pe_diff}, {gm_diff} margin |\n")
            f.write(f"| **{sym}** | **{q['pe']:.0f}x** | **{fin['gm']*100:.0f}%** | Subject |\n\n")

        f.write("---\n\n")
        f.write("## 8. Investment Thesis\n\n")
        f.write("### Strengths\n")

        if sym == "META":
            f.write("- Exceptional 81% gross margin, industry-leading\n")
            f.write("- AI-driven ad targeting improving ROI\n")
            f.write("- Reels monetization gaining traction\n")
            f.write("- WhatsApp monetization potential\n")
        elif sym == "NVDA":
            f.write("- Dominant GPU market share in AI/ML\n")
            f.write("- CUDA ecosystem moat\n")
            f.write("- Exceptional 73% gross margin\n")
            f.write("- Data center revenue accelerating\n")
        elif sym == "MSFT":
            f.write("- Cloud leader with Azure growth\n")
            f.write("- Enterprise software lock-in\n")
            f.write("- AI integration across products\n")
        elif sym == "AAPL":
            f.write("- Premium brand moat\n")
            f.write("- Services revenue growth\n")
            f.write("- Strong ecosystem lock-in\n")
        elif sym == "JPM":
            f.write("- Market leader in investment banking\n")
            f.write("- Diversified revenue streams\n")
            f.write("- Strong risk management\n")
        elif sym == "JNJ":
            f.write("- Diversified healthcare portfolio\n")
            f.write("- Strong dividend growth\n")
            f.write("- Pipeline of new drugs\n")
        f.write("\n### Concerns\n")

        if sym == "META":
            f.write("- Regulatory scrutiny\n")
            f.write("- Heavy AI capex spending\n")
            f.write("- Competition from TikTok\n")
        elif sym == "NVDA":
            f.write("- Valuation extremely demanding\n")
            f.write("- Competition from AMD/custom silicon\n")
            f.write("- Cyclical semiconductor risk\n")
        elif sym == "MSFT":
            f.write("- Valuation at historical highs\n")
            f.write("- Regulatory scrutiny\n")
        elif sym == "AAPL":
            f.write("- China revenue dependence\n")
            f.write("- Mature smartphone market\n")
        elif sym == "JPM":
            f.write("- Interest rate sensitivity\n")
            f.write("- Credit cycle risk\n")
        f.write("\n### Catalysts\n")

        if sym == "META":
            f.write("- Meta AI monetization acceleration\n")
            f.write("- WhatsApp Business revenue scaling\n")
            f.write("- Reels ad revenue growth\n")
        elif sym == "NVDA":
            f.write("- Blackwell GPU ramp\n")
            f.write("- AI infrastructure spending continues\n")
        elif sym == "MSFT":
            f.write("- Copilot adoption\n")
            f.write("- Azure AI services growth\n")
        f.write("\n")
        f.write("---\n\n")
        f.write("## 5. Recommendation\n\n")
        f.write(f"**Rating: {rating_str}** | Score: {sc}/100 | Risk: {risk}\n\n")

        if upside > 30:
            f.write("**STRONG BUY**: Significant upside potential based on analyst consensus and DCF valuation.\n\n")
        elif upside > 15:
            f.write("**BUY**: Moderate upside with solid fundamentals.\n\n")
        elif upside > 0:
            f.write("**HOLD**: Limited upside, wait for better entry.\n\n")
        else:
            f.write("**SELL**: Downside risk exceeds upside potential.\n\n")

        # Fair Value Range
        f.write("\n### Fair Value Range\n\n")
        f.write(f"| Level | Price | vs Current |\n")
        f.write(f"|-------|-------|------------|\n")
        f.write(f"| Bear Case | ${fv_low:.2f} | {(fv_low-price)/price*100:+.1f}% |\n")
        f.write(f"| **Base Case** | **${target:.2f}** | **{upside:+.1f}%** |\n")
        f.write(f"| Bull Case | ${fv_high:.2f} | {(fv_high-price)/price*100:+.1f}% |\n\n")

        # Key metrics summary
        f.write("\n### Key Metrics\n\n")
        f.write(f"| Metric | Value |\n")
        f.write(f"|--------|-------|\n")
        f.write(f"| P/E | {q['pe']:.0f}x |\n")
        f.write(f"| PEG | {q['peg']:.2f} |\n")
        f.write(f"| FCF Yield | {cf['fcf_yield']*100:.1f}% |\n")
        f.write(f"| ROE | {fin['roe']*100:.0f}% |\n\n")

        f.write("---\n\n")
        f.write("*Generated by Stock PRO v8.0*\n")

    print(f"\nReport saved: {md_path}")
    return md_path

def generate_report_cn(sym):
    """Generate Chinese stock analysis report"""
    sym = sym.upper().strip()
    print(f"\n{'='*60}\nStock PRO v9.0 - 分析 {sym}\n{'='*60}")

    q = fetch(sym)
    if not q.get("success"):
        print(f"获取数据失败: {q.get('error', '未知错误')}")
        return None

    fin = FD.get(sym, FD["AAPL"])
    analyst = ANALYST.get(sym, {"target": q["price"], "rating": "N/A"})

    price = q["price"]
    analyst_target = analyst["target"]

    dcf_scenarios = calc_dcf_scenarios(q, fin)
    dcf_base = dcf_scenarios["Base"]
    target = analyst_target * 0.6 + dcf_base * 0.4
    upside = (target - price) / price * 100

    sc = score_stock(q, fin)
    risk = risk_level(q, fin)
    pe_f = price / q["eps_f"]

    eps_proj = calc_5y_eps(q, fin)
    risk_m = calc_risk_metrics(q)
    tech = calc_technical(q)
    val_pct = calc_valuation_percentile(q, fin)
    cf = calc_cash_flow_analysis(q, fin)
    mom = calc_momentum(q)
    traps = calc_value_trap_detection(q, fin)
    fv_low, fv_high = calc_fair_value_range(dcf_scenarios, analyst_target)
    sr_levels = calc_support_resistance(q)
    analyst_ratings = calc_analyst_consensus(sym)
    inst_own = calc_institutional_ownership(sym)

    quality = calc_quality_score(q, fin)
    value = calc_value_score(q, fin, val_pct)
    momentum_s = calc_momentum_score(q, mom)
    growth = calc_growth_score(q, fin)
    sc = int(quality * 0.30 + value * 0.25 + momentum_s * 0.20 + growth * 0.25)

    # New analysis modules
    earnings_cal = calc_earnings_calendar(sym)
    insider = calc_insider_sentiment(sym)
    tech_ind = calc_technical_indicators(q, fin)
    sector = calc_sector_comparison(sym, q, fin)
    div_info = calc_dividend_analysis(q, fin, cf)
    prob_target = calc_price_target_probability(price, target, fv_low, fv_high, dcf_base)
    balance = calc_balance_sheet_health(fin, q)
    competitors = calc_competitors(sym)
    cap_class, cap_range = calc_market_cap_class(q["mc"])
    price_chart = calc_price_chart(q)
    forecast = calc_fundamental_forecast(q, fin)
    money_flow = calc_money_flow(q)
    swot = calc_swot_analysis(sym, q, fin, val_pct, cf)
    comp_data = calc_competitor_comparison(sym, q, fin)

    rating_cn = {"STRONG_BUY": "强烈买入", "BUY": "买入", "HOLD": "持有", "SELL": "卖出", "STRONG_SELL": "强烈卖出"}.get(rating(upside), rating(upside))
    risk_cn = {"LOW": "低风险", "MEDIUM": "中等风险", "HIGH": "高风险"}.get(risk, risk)

    print(f"\n价格: ${price:.2f}")
    print(f"EPS: 追踪 ${q['eps_t']:.2f} | 预测 ${q['eps_f']:.2f}")
    print(f"P/E: 当前 {q['pe']:.1f}x | 预测 {pe_f:.1f}x")
    print(f"\nDCF情景: 乐观=${dcf_scenarios['Bull']:.0f} | 基准=${dcf_base:.0f} | 悲观=${dcf_scenarios['Bear']:.0f}")
    print(f"分析师目标: ${analyst_target:.0f} ({analyst['rating']})")
    print(f"综合目标: ${target:.2f} ({upside:+.1f}%)")
    print(f"\n评级: {rating_cn} | 得分: {sc}/100 | 风险: {risk_cn}")
    print(f"得分: 质量={quality} | 价值={value} | 动量={momentum_s} | 成长={growth}")
    print(f"资金流向: {money_flow['flow']} | SWOT: {len(swot['strengths'])}S/{len(swot['weaknesses'])}W")

    date = datetime.now().strftime("%Y-%m-%d")
    md_path = Path(f"50-reports/stocks/{sym}_{date}.md")
    md_path.parent.mkdir(parents=True, exist_ok=True)

    with open(md_path, "w", encoding="utf-8") as f:
        # Header
        rating_icon = {"STRONG_BUY": "🟢", "BUY": "🟢", "HOLD": "🟡", "SELL": "🔴", "STRONG_SELL": "🔴"}.get(rating(upside), "⚪")
        f.write(f"# {sym} 股票分析报告\n\n")
        f.write(f"**生成日期:** {date} | **版本:** v8.0\n\n")
        f.write("---\n\n")
        f.write(f"## {rating_icon} 投资评级: {rating_cn}\n\n")
        f.write(f"| 关键指标 | 数值 |\n")
        f.write(f"|---------|------|\n")
        f.write(f"| 当前价格 | **${price:.2f}** |\n")
        f.write(f"| 12个月目标 | **${target:.2f}** |\n")
        f.write(f"| 上涨空间 | **{upside:+.1f}%** |\n")
        f.write(f"| 综合评分 | **{sc}/100** |\n")
        f.write(f"| 风险等级 | {risk_cn} |\n\n")
        f.write("---\n\n")

        # Score Breakdown
        f.write("## 1. 执行摘要\n\n")
        f.write(f"| 指标 | 数值 | 评估 |\n")
        f.write(f"|------|------|------|\n")
        f.write(f"| 当前价格 | ${price:.2f} | - |\n")
        f.write(f"| 分析师目标 | ${analyst_target:.0f} | {analyst['rating']} |\n")
        f.write(f"| DCF基准 | ${dcf_base:.2f} | 内在价值 |\n")
        f.write(f"| 综合目标 | ${target:.2f} | **最佳估算** |\n")
        f.write(f"| 上涨/下跌 | {upside:+.1f}% | {rating_cn} |\n")
        f.write(f"| 风险 | {risk_cn} | - |\n\n")

        f.write("### 1.1 评分明细\n\n")
        f.write(f"| 维度 | 得分 | 权重 | 贡献 |\n")
        f.write(f"|------|------|------|------|\n")
        f.write(f"| 质量 | {quality} | 30% | {quality*0.30:.0f} |\n")
        f.write(f"| 价值 | {value} | 25% | {value*0.25:.0f} |\n")
        f.write(f"| 成长 | {growth} | 25% | {growth*0.25:.0f} |\n")
        f.write(f"| 动量 | {momentum_s} | 20% | {momentum_s*0.20:.0f} |\n")
        f.write(f"| **综合** | **{sc}** | 100% | - |\n\n")

        score_bar = "█" * int(sc/5) + "░" * (20 - int(sc/5))
        f.write(f"综合评分: [{score_bar}] {sc}/100\n\n")
        f.write("---\n\n")

        # Valuation
        f.write("## 2. 估值分析\n\n")
        f.write(f"| 指标 | 数值 | 评估 |\n")
        f.write(f"|------|------|------|\n")
        pe_assess = "偏高" if q["pe"] > 30 else "合理" if q["pe"] > 15 else "偏低"
        f.write(f"| 市盈率 (TTM) | {q['pe']:.1f}x | {pe_assess} |\n")
        pe_f_assess = "偏高" if pe_f > 25 else "合理" if pe_f > 15 else "偏低"
        f.write(f"| 市盈率 (预测) | {pe_f:.1f}x | {pe_f_assess} |\n")
        peg_assess = "成长性强" if q["peg"] < 1 else "合理" if q["peg"] < 2 else "偏高"
        f.write(f"| PEG | {q['peg']:.2f} | {peg_assess} |\n")
        f.write(f"| EV/EBITDA | {fin['ev_ebitda']:.1f}x | - |\n\n")

        f.write("### 2.1 DCF三情景分析\n\n")
        f.write(f"| 情景 | 目标价 | 上涨空间 |\n")
        f.write(f"|------|--------|----------|\n")
        f.write(f"| 乐观 (Bull) | ${dcf_scenarios['Bull']:.0f} | +{(dcf_scenarios['Bull']-price)/price*100:.1f}% |\n")
        f.write(f"| 基准 (Base) | ${dcf_base:.0f} | +{(dcf_base-price)/price*100:.1f}% |\n")
        f.write(f"| 悲观 (Bear) | ${dcf_scenarios['Bear']:.0f} | +{(dcf_scenarios['Bear']-price)/price*100:.1f}% |\n\n")

        f.write("### 2.2 估值定位\n\n")
        f.write(f"| 分位 | 评估 |\n")
        f.write(f"|------|------|\n")
        f.write(f"| P/E历史分位: **{val_pct['percentile']:.0f}%** | {val_pct['assessment']} |\n\n")

        # P/E position bar
        pct_bar = "█" * int(val_pct['percentile']/5) + "░" * (20 - int(val_pct['percentile']/5))
        f.write(f"P/E位置: [{pct_bar}] {val_pct['percentile']:.0f}%\n\n")

        f.write("---\n\n")

        # EPS Projection
        f.write("## 3. 盈利预测 (5年)\n\n")
        f.write(f"| 年份 | EPS | 增长率 | 说明 |\n")
        f.write(f"|------|-----|--------|------|\n")
        for proj in eps_proj:
            note = "高增长" if proj["growth"] > 20 else "稳定" if proj["growth"] > 10 else "成熟期"
            f.write(f"| Y{proj['year']} | ${proj['eps']:.2f} | {proj['growth']:.1f}% | {note} |\n")
        f.write(f"| **CAGR** | - | **{eps_proj[-1]['growth']:.1f}%** | 5年复合 |\n\n")
        f.write("---\n\n")

        # Risk
        f.write("## 4. 风险分析\n\n")

        f.write("### 4.1 技术指标\n\n")
        f.write(f"| 指标 | 数值 |\n")
        f.write(f"|------|------|\n")
        f.write(f"| 趋势 | {tech['trend']} |\n")
        f.write(f"| 距52周高点 | {tech['dist_from_high']:.1f}% |\n")
        f.write(f"| 距52周低点 | +{tech['dist_from_low']:.1f}% |\n")
        f.write(f"| 均线位置 | {tech['ma_status']} |\n\n")

        f.write("### 4.2 风险指标\n\n")
        f.write(f"| 指标 | 数值 | 说明 |\n")
        f.write(f"|------|------|------|\n")
        f.write(f"| VaR (95%) | ${risk_m['var_95']:.2f} ({risk_m['var_95_pct']:.1f}%) | 单日最大损失 |\n")
        f.write(f"| 最大回撤风险 | {risk_m['max_dd_pct']:.1f}% | 从52周高点 |\n")
        f.write(f"| 夏普比率 | {risk_m['sharpe']:.2f} | 风险调整收益 |\n")
        f.write(f"| 52周位置 | {risk_m['range_pos']:.0f}% | 距低点百分比 |\n\n")
        f.write(f"| 52周区间 | ${q['w52l']:.0f} - ${q['w52h']:.0f} | 当前: ${price:.2f} |\n\n")

        f.write("### 4.3 支撑与阻力\n\n")

        f.write("### 4.2 支撑与阻力\n\n")
        f.write(f"| 级别 | 价格 | 距当前 |\n")
        f.write(f"|------|------|--------|\n")
        f.write(f"| 阻力 R4 | ${sr_levels['r4']:.2f} | +{(sr_levels['r4']-price)/price*100:.1f}% |\n")
        f.write(f"| 阻力 R3 | ${sr_levels['r3']:.2f} | +{(sr_levels['r3']-price)/price*100:.1f}% |\n")
        f.write(f"| 阻力 R2 | ${sr_levels['r2']:.2f} | +{(sr_levels['r2']-price)/price*100:.1f}% |\n")
        f.write(f"| 阻力 R1 | ${sr_levels['r1']:.2f} | +{(sr_levels['r1']-price)/price*100:.1f}% |\n")
        f.write(f"| **当前价格** | **${price:.2f}** | 0.0% |\n")
        f.write(f"| 支撑 S1 | ${sr_levels['s1']:.2f} | {(sr_levels['s1']-price)/price*100:.1f}% |\n")
        f.write(f"| 支撑 S2 | ${sr_levels['s2']:.2f} | {(sr_levels['s2']-price)/price*100:.1f}% |\n")
        f.write(f"| 支撑 S3 | ${sr_levels['s3']:.2f} | {(sr_levels['s3']-price)/price*100:.1f}% |\n\n")
        f.write("---\n\n")

        # Financial Health
        f.write("## 5. 财务健康\n\n")
        f.write(f"| 指标 | 数值 | 评级 |\n")
        f.write(f"|------|------|------|\n")
        gm_assess = "优秀" if fin["gm"] > 0.40 else "良好" if fin["gm"] > 0.25 else "一般"
        f.write(f"| 毛利率 | {fin['gm']*100:.1f}% | {gm_assess} |\n")
        nm_assess = "优秀" if fin["nm"] > 0.20 else "良好" if fin["nm"] > 0.10 else "一般"
        f.write(f"| 净利率 | {fin['nm']*100:.1f}% | {nm_assess} |\n")
        roe_assess = "优秀" if fin["roe"] > 0.20 else "良好" if fin["roe"] > 0.10 else "一般"
        f.write(f"| ROE | {fin['roe']*100:.1f}% | {roe_assess} |\n")
        roic_assess = "优秀" if fin["roic"] > 0.15 else "良好" if fin["roic"] > 0.08 else "一般"
        f.write(f"| ROIC | {fin['roic']*100:.1f}% | {roic_assess} |\n")
        f.write(f"| 负债率 | {fin['de']:.1f}x | {'低' if fin['de'] < 1 else '中等' if fin['de'] < 2 else '高'} |\n")
        f.write(f"| Beta | {q['beta']:.2f} | {'高波动' if q['beta'] > 1.5 else '中等' if q['beta'] > 0.8 else '低波动'} |\n\n")

        f.write("### 5.1 现金流分析\n\n")
        f.write(f"| 指标 | 数值 | 评估 |\n")
        f.write(f"|------|------|------|\n")
        fcf_assess = "良好" if cf['fcf_yield'] > 0.04 else "一般" if cf['fcf_yield'] > 0.02 else "较低"
        f.write(f"| FCF收益率 | {cf['fcf_yield']*100:.1f}% | {fcf_assess} |\n")
        div_assess = "有" if cf['div_yield'] > 0.02 else "无/低"
        f.write(f"| 股息率 | {cf['div_yield']*100:.2f}% | {div_assess} |\n")
        f.write(f"| 派息率 | {fin['payout']*100:.0f}% | {'可持续' if fin['payout'] < 0.6 else '较高'} |\n\n")

        # Balance Sheet Health
        f.write("### 5.2 资产负债表健康度\n\n")
        f.write(f"| 指标 | 数值 | 状态 |\n")
        f.write(f"|------|------|------|\n")
        z_color_icon = {"Green": "✓", "Yellow": "⚠", "Red": "✗"}.get(balance['z_color'], "")
        z_status_cn = {"Safe Zone": "安全区", "Grey Zone": "灰色区", "Distress Zone": "困境区"}.get(balance['z_status'], balance['z_status'])
        f.write(f"| Altman Z评分 | {balance['z_score']:.1f} | {z_color_icon} {z_status_cn} |\n")
        f.write(f"| 流动比率 | {balance['current_ratio']:.1f}x | {'良好' if balance['current_ratio'] > 1.5 else '一般'} |\n")
        f.write(f"| 负债能力 | {balance['debt_capacity']} | 空间: {balance['headroom_pct']}% |\n")
        f.write(f"| 市值 | ${q['mc']/1e9:.0f}B | {cap_class} |\n\n")

        # 52W Price Chart
        f.write("### 5.3 价格位置图\n\n")
        f.write(f"```\n")
        f.write(f"52周高点 ${q['w52h']:.0f} ┬\n")
        chart_line = "                    " + "█" * int(price_chart['position']) + "●\n"
        f.write(f"{chart_line}")
        f.write(f"                    {price_chart['chart']}\n")
        f.write(f"52周低点 ${q['w52l']:.0f} ┴\n")
        f.write(f"```\n")
        f.write(f"当前价格: ${price:.2f} (区间内{(price - q['w52l'])/(q['w52h'] - q['w52l'])*100:.0f}%位置)\n\n")

        f.write("---\n\n")

        # Analyst
        f.write("## 6. 分析师观点\n\n")
        f.write(f"| 评级 | 占比 | 说明 |\n")
        f.write(f"|------|------|------|\n")
        f.write(f"| 买入/增持 | {analyst_ratings['buy']}% | 强烈看好 |\n")
        f.write(f"| 持有 | {analyst_ratings['hold']}% | 中性 |\n")
        f.write(f"| 卖出/减持 | {analyst_ratings['sell']}% | 负面 |\n\n")

        buy_bar = "█" * int(analyst_ratings['buy'] / 5)
        hold_bar = "█" * int(analyst_ratings['hold'] / 5)
        sell_bar = "█" * int(analyst_ratings['sell'] / 5)
        f.write(f"评级分布: [{buy_bar}]({analyst_ratings['buy']}% 买入) [{hold_bar}]({analyst_ratings['hold']}% 持有) [{sell_bar}]({analyst_ratings['sell']}% 卖出)\n\n")

        f.write(f"| 指标 | 数值 |\n")
        f.write(f"|------|------|\n")
        f.write(f"| 覆盖分析师 | {analyst.get('num', 'N/A')} | - |\n")
        f.write(f"| 一致评级 | {analyst['rating']} | - |\n")
        f.write(f"| 12个月目标 | ${analyst_target:.0f} | {upside:+.1f}% 空间 |\n\n")

        f.write("### 6.1 机构持仓\n\n")
        f.write(f"| 指标 | 数值 | 说明 |\n")
        f.write(f"|------|------|------|\n")
        inst_assess = "高 (稳定)" if inst_own > 65 else "中等" if inst_own > 50 else "低 (波动)"
        f.write(f"| 机构持仓比例 | {inst_own}% | {inst_assess} |\n\n")

        # Insider Trading
        f.write("### 6.2 内部人交易\n\n")
        f.write(f"| 指标 | 数值 |\n")
        f.write(f"|------|------|\n")
        f.write(f"| 30日买入 | {insider['buy_30d']} |\n")
        f.write(f"| 30日卖出 | {insider['sell_30d']} |\n")
        f.write(f"| 情绪 | {insider['sentiment']} |\n\n")

        # Earnings Calendar
        f.write("### 6.3 财报日历\n\n")
        f.write(f"| 指标 | 数值 |\n")
        f.write(f"|------|------|\n")
        f.write(f"| 财报季 | {earnings_cal['month']} |\n")
        f.write(f"| 影响程度 | {earnings_cal['impact']} |\n")
        f.write(f"| 提前关注 | -{earnings_cal['days_before']}天 |\n\n")

        # Technical Indicators
        f.write("### 6.4 技术指标\n\n")
        f.write(f"| 指标 | 数值 | 信号 |\n")
        f.write(f"|------|------|------|\n")
        f.write(f"| RSI (14) | {tech_ind['rsi']:.0f} | {tech_ind['rsi_signal']} |\n")
        f.write(f"| MACD | - | {tech_ind['macd']} |\n")
        f.write(f"| MA50 | ${tech_ind['ma_50']:.2f} | Price {tech_ind['price_vs_ma50']} |\n")
        f.write(f"| MA200 | ${tech_ind['ma_200']:.2f} | Price {tech_ind['price_vs_ma200']} |\n\n")

        # Sector Comparison
        f.write(f"### 6.5 行业对比 ({sector['sector']})\n\n")
        f.write(f"| 指标 | 股票 | 行业平均 | 对比 |\n")
        f.write(f"|------|------|----------|------|\n")
        f.write(f"| P/E | {sector['stock_pe']:.1f}x | {sector['avg_pe']:.1f}x | {sector['pe_vs']} |\n")
        f.write(f"| PEG | {sector['stock_peg']:.2f} | {sector['avg_peg']:.2f} | {sector['peg_vs']} |\n")
        f.write(f"| ROE | {fin['roe']*100:.1f}% | {sector['roe_vs']} | - |\n")
        f.write(f"| FCF Yield | {fin.get('fcf_yield', 0)*100:.1f}% | {sector['fcf_vs']} | - |\n\n")

        # Dividend
        f.write("### 6.6 分红分析\n\n")
        f.write(f"| 指标 | 数值 |\n")
        f.write(f"|------|------|\n")
        f.write(f"| 股息率 | {div_info['div_yield']*100:.2f}% |\n")
        f.write(f"| 派息率 | {div_info['payout']*100:.0f}% |\n")
        f.write(f"| 状态 | {div_info['status']} |\n")
        f.write(f"| 健康度 | {div_info['health']} |\n\n")

        # Money Flow
        f.write("### 5.4 资金流向分析\n\n")
        f.write(f"| 指标 | 数值 | 信号 |\n")
        f.write(f"|------|------|------|\n")
        flow_icon = {"Bullish": "↑", "Bearish": "↓", "Neutral": "→"}.get(money_flow['signal'], "")
        f.write(f"| 资金流向 | {money_flow['flow']} | {flow_icon} {money_flow['signal']} |\n")
        f.write(f"| 积累/分配 | {money_flow['accum_dist']} | - |\n")
        f.write(f"| 成交量权重 | {money_flow['vol_weight']} | - |\n\n")

        # Fundamental Forecast
        f.write("### 5.5 盈利预测模型\n\n")
        f.write(f"| P/E情景 | 目标价 | 上涨空间 |\n")
        f.write(f"|---------|--------|----------|\n")
        f.write(f"| P/E 15x (保守) | ${forecast['pe_15']:.0f} | {(forecast['pe_15']-price)/price*100:+.1f}% |\n")
        f.write(f"| P/E 20x (平均) | ${forecast['pe_20']:.0f} | {(forecast['pe_20']-price)/price*100:+.1f}% |\n")
        f.write(f"| P/E 25x (乐观) | ${forecast['pe_25']:.0f} | {(forecast['pe_25']-price)/price*100:+.1f}% |\n")
        f.write(f"| P/E 30x (激进) | ${forecast['pe_30']:.0f} | {(forecast['pe_30']-price)/price*100:+.1f}% |\n\n")
        f.write(f"| 增长率 | EPS: {forecast['eps_growth']:.1f}% | 营收: {forecast['rev_growth']:.1f}% |\n")
        f.write(f"| 内在价值 | ${forecast['intrinsic']:.0f} | {(forecast['intrinsic']-price)/price*100:+.1f}% |\n\n")

        # SWOT Analysis
        f.write("### 5.6 SWOT分析\n\n")
        f.write(f"| 优势 (S) | 劣势 (W) |\n")
        f.write(f"|----------|----------|\n")
        s_len = len(swot['strengths'])
        w_len = len(swot['weaknesses'])
        for i in range(max(s_len, w_len)):
            s = swot['strengths'][i] if i < s_len else ""
            w = swot['weaknesses'][i] if i < w_len else ""
            f.write(f"| {s} | {w} |\n")
        f.write(f"\n| 机会 (O) | 威胁 (T) |\n")
        f.write(f"|----------|----------|\n")
        o_len = len(swot['opportunities'])
        t_len = len(swot['threats'])
        for i in range(max(o_len, t_len)):
            o = swot['opportunities'][i] if i < o_len else ""
            t = swot['threats'][i] if i < t_len else ""
            f.write(f"| {o} | {t} |\n")
        f.write("\n")

        f.write("---\n\n")

        # Investment Recommendation
        f.write("## 7. 投资建议\n\n")

        # Probability-weighted target
        f.write("### 7.1 概率加权目标价\n\n")
        f.write(f"| 情景 | 概率 | 目标价 | 上涨空间 |\n")
        f.write(f"|------|------|--------|----------|\n")
        f.write(f"| 悲观 (15%) | 15% | ${prob_target['conservative']:.0f} | {(prob_target['conservative']-price)/price*100:+.1f}% |\n")
        f.write(f"| 基准 (50%) | 50% | ${prob_target['base']:.0f} | {upside:+.1f}% |\n")
        f.write(f"| 乐观 (25%) | 25% | ${prob_target['optimistic']:.0f} | {(prob_target['optimistic']-price)/price*100:+.1f}% |\n\n")
        f.write(f"**期望目标价:** ${prob_target['expected']:.0f} ({prob_target['upside_pct']:+.1f}%)\n\n")
        f.write(f"| 情景 | 目标价 | 上涨空间 | 概率 |\n")
        f.write(f"|------|--------|----------|------|\n")
        f.write(f"| 悲观 | ${fv_low:.0f} | {(fv_low-price)/price*100:+.1f}% | 15% |\n")
        f.write(f"| 基准 | ${target:.0f} | {upside:+.1f}% | 60% |\n")
        f.write(f"| 乐观 | ${fv_high:.0f} | {(fv_high-price)/price*100:+.1f}% | 25% |\n\n")

        f.write(f"**合理价值区间:** ${fv_low:.0f} - ${fv_high:.0f}\n\n")

        if upside > 25:
            f.write(f"**{rating_cn}** - 目标上涨空间 {upside:+.1f}%，具备显著投资价值\n\n")
        elif upside > 10:
            f.write(f"**{rating_cn}** - 目标上涨空间 {upside:+.1f}%，建议适度配置\n\n")
        else:
            f.write(f"**{rating_cn}** - 目标上涨空间有限，建议观望\n\n")

        # Value traps
        if traps:
            f.write("### 7.1 风险提示\n\n")
            for trap in traps:
                f.write(f"- ⚠️ {trap['type']}: {trap['detail']}\n")
            f.write("\n")

        f.write("---\n\n")
        f.write(f"*由 Stock PRO v8.0 生成*\n")

    print(f"\n报告已保存: {md_path}")
    return md_path

def calc_fundamental_forecast(q, fin):
    """Calculate fundamental-based price targets"""
    price = q["price"]
    eps_t = q["eps_t"]
    eps_f = q["eps_f"]
    pe = q["pe"]

    # Earnings-based targets
    pe_15 = eps_f * 15  # Conservative P/E
    pe_20 = eps_f * 20  # Average P/E
    pe_25 = eps_f * 25  # Bullish P/E
    pe_30 = eps_f * 30  # Aggressive P/E

    # Growth-based
    rev_g = fin.get("rev_g", 0.10)
    eps_g = (eps_f - eps_t) / eps_t if eps_t > 0 else 0.10
    eps_cagr = eps_g  # Simplified

    # Intrinsic value estimate (DCF-like)
    if eps_f > 0:
        # Gordon Growth Model approximation
        discount_rate = 0.10
        growth_rate = min(eps_g, 0.25)  # Cap at 25%
        intrinsic = eps_f * (1 + growth_rate) / (discount_rate - growth_rate) if growth_rate < discount_rate else eps_f * 10
    else:
        intrinsic = price

    return {
        "pe_15": pe_15,
        "pe_20": pe_20,
        "pe_25": pe_25,
        "pe_30": pe_30,
        "intrinsic": intrinsic,
        "eps_growth": eps_g * 100,
        "rev_growth": rev_g * 100,
    }

def calc_money_flow(q):
    """Estimate money flow indicators"""
    price = q["price"]
    w52h = q["w52h"]
    w52l = q["w52l"]

    # Volume estimation (relative)
    range_pos = (price - w52l) / (w52h - w52l) if w52h > w52l else 0.5

    # Money flow direction
    if range_pos > 0.6:
        flow = "Institutional Inflow"
        signal = "Bullish"
    elif range_pos < 0.4:
        flow = "Institutional Outflow"
        signal = "Bearish"
    else:
        flow = "Neutral"
        signal = "Neutral"

    # Price relative to volume
    vol_weight = "High" if range_pos > 0.7 or range_pos < 0.3 else "Normal"

    return {
        "flow": flow,
        "signal": signal,
        "vol_weight": vol_weight,
        "accum_dist": "Accumulating" if range_pos < 0.4 else "Distributing",
    }

def calc_swot_analysis(sym, q, fin, val_pct, cf):
    """Generate SWOT analysis"""
    strengths = []
    weaknesses = []
    opportunities = []
    threats = []

    # Strengths
    if fin["roe"] > 0.25:
        strengths.append("高ROE (>25%)")
    if fin["gm"] > 0.40:
        strengths.append("高毛利率")
    if fin.get("fcf_yield", 0) > 0.04:
        strengths.append("良好现金流")
    if fin["de"] < 0.5:
        strengths.append("低负债")
    if q["beta"] < 1.0:
        strengths.append("低波动性")

    # Weaknesses
    if fin["nm"] < 0.10:
        weaknesses.append("低净利率")
    if fin["de"] > 1.5:
        weaknesses.append("高负债水平")
    if val_pct["percentile"] > 70:
        weaknesses.append("估值偏高")
    if q["peg"] > 2.0:
        weaknesses.append("PEG偏高")
    if fin.get("fcf_yield", 0) < 0:
        weaknesses.append("负现金流")

    # Opportunities
    if fin.get("rev_g", 0) > 0.20:
        opportunities.append("高营收增长")
    if q["eps_f"] > q["eps_t"] * 1.15:
        opportunities.append("盈利增长加速")
    if val_pct["percentile"] < 40:
        opportunities.append("估值处于低位")

    # Threats
    if q["beta"] > 1.5:
        threats.append("高波动性")
    if fin["de"] > 2.0:
        threats.append("债务风险")
    if fin.get("rev_g", 0) < 0.05:
        threats.append("增长放缓")

    return {
        "strengths": strengths[:5],
        "weaknesses": weaknesses[:5],
        "opportunities": opportunities[:5],
        "threats": threats[:5],
    }

def calc_competitor_comparison(sym, q, fin):
    """Compare with key competitors"""
    COMP_DATA = {
        "META": [
            {"sym": "GOOGL", "pe": 25, "peg": 1.2, "roe": 0.28, "fcf": 0.045},
            {"sym": "SNAP", "pe": 0, "peg": 0, "roe": -0.15, "fcf": -0.02},
        ],
        "NVDA": [
            {"sym": "AMD", "pe": 45, "peg": 1.5, "roe": 0.10, "fcf": 0.02},
            {"sym": "INTC", "pe": 20, "peg": 3.0, "roe": 0.08, "fcf": 0.04},
        ],
        "AAPL": [
            {"sym": "MSFT", "pe": 36, "peg": 2.5, "roe": 0.38, "fcf": 0.03},
            {"sym": "GOOGL", "pe": 25, "peg": 1.2, "roe": 0.28, "fcf": 0.045},
        ],
        "TSLA": [
            {"sym": "F", "pe": 8, "peg": 1.0, "roe": 0.05, "fcf": 0.03},
            {"sym": "GM", "pe": 6, "peg": 1.2, "roe": 0.08, "fcf": 0.02},
        ],
    }

    return COMP_DATA.get(sym, [])

def generate_comparison(symbols):
    """Generate comparison table for multiple stocks"""
    print(f"\n{'='*80}")
    print(f"Stock PRO v10.0 - Multi-Stock Comparison")
    print(f"{'='*80}\n")

    results = []
    for sym in symbols:
        sym = sym.upper().strip()
        q = fetch(sym)
        if not q.get("success"):
            print(f"Failed to fetch {sym}: {q.get('error', 'Unknown')}")
            continue

        fin = FD.get(sym, FD["AAPL"])
        analyst = ANALYST.get(sym, {"target": q["price"], "rating": "N/A"})

        price = q["price"]
        analyst_target = analyst["target"]
        dcf_scenarios = calc_dcf_scenarios(q, fin)
        dcf_base = dcf_scenarios["Base"]
        target = analyst_target * 0.6 + dcf_base * 0.4
        upside = (target - price) / price * 100

        quality = calc_quality_score(q, fin)
        value = calc_value_score(q, fin, {"percentile": 50})
        momentum_s = calc_momentum_score(q, {"range_pos": 50, "momentum": "Neutral"})
        growth = calc_growth_score(q, fin)
        sc = int(quality * 0.30 + value * 0.25 + momentum_s * 0.20 + growth * 0.25)

        results.append({
            "symbol": sym,
            "price": price,
            "target": target,
            "upside": upside,
            "score": sc,
            "pe": q["pe"],
            "peg": q["peg"],
            "fcf_yield": fin.get("fcf_yield", 0),
            "div_yield": fin.get("div_yield", 0),
            "roe": fin["roe"],
            "beta": q["beta"],
            "rating": rating(upside),
        })

    # Sort by upside
    results.sort(key=lambda x: x["upside"], reverse=True)

    # Print comparison table
    print(f"{'Symbol':<8} {'Price':>8} {'Target':>8} {'Upside':>8} {'Score':>6} {'P/E':>6} {'PEG':>5} {'FCF':>6} {'Div':>5} {'ROE':>6} {'Rating'}")
    print("-" * 100)

    for r in results:
        rating_icon = {"STRONG_BUY": "STRONG_BUY", "BUY": "BUY", "HOLD": "HOLD", "SELL": "SELL", "STRONG_SELL": "STRONG_SELL"}.get(r["rating"], "")
        print(f"{r['symbol']:<8} ${r['price']:>7.2f} ${r['target']:>7.2f} {r['upside']:>+7.1f}% {r['score']:>5}/100 {r['pe']:>5.1f}x {r['peg']:>4.2f} {r['fcf_yield']*100:>5.1f}% {r['div_yield']*100:>4.2f}% {r['roe']*100:>5.1f}% {rating_icon}")

    print("-" * 100)
    print(f"\nTotal: {len(results)} stocks analyzed")
    return results

def print_banner():
    """Print banner"""
    print("""
+==============================================================+
|                    Stock PRO v10.0                         |
|              Professional Stock Analysis Tool               |
+==============================================================+
    """)

def print_help():
    """Print help menu"""
    print("""
Stock PRO v10.0 - Professional Stock Analysis Tool
=====================================================

Usage:
   py stock_pro_v4.py <SYMBOL> [OPTIONS]

Analysis Modes:
   NVDA             English report
   NVDA --cn        Chinese report
   NVDA --json      JSON output
   NVDA --chart     Generate chart
   NVDA --news      Show news

Multi-Stock:
   --compare SYM...  Compare multiple stocks
   --summary SYM...  Summary cards

Data Export:
   --csv SYM...      Export to CSV
   --xlsx SYM...     Export to Excel
   --db SYM...       Save to SQLite

Automation:
   --watch SYM...    Auto-refresh (30s)
   --alert SYM...    Price alerts
   --cron "expr"    Setup cron job
   --notify SYM...   Send notifications

Portfolio:
   --portfolio        View portfolio
   --portfolio-add SYM SHARES COST  Add position
   --portfolio-remove SYM          Remove position

Screener:
   --screener              Show all qualifying stocks
   --screener --min-score 80  Custom score threshold
   --screener --min-upside 20 Custom upside threshold

Email/Notifications:
   --email --to user@email.com  Send report
   --webhook --url URL          Send webhook

Advanced:
   --api          Start REST API server
   --dashboard    Create HTML dashboard
   --pipeline     Create batch pipeline
   --skill        Create CoPaw skill
   --config       Show config
   --news SYM...  Show news

Examples:
   py stock_pro_v4.py NVDA                    # Single stock
   py stock_pro_v4.py --compare NVDA META    # Compare
   py stock_pro_v4.py --screener             # Stock screener
   py stock_pro_v4.py --portfolio            # View portfolio
   py stock_pro_v4.py --portfolio-add NVDA 100 170
   py stock_pro_v4.py --news NVDA META       # News
   py stock_pro_v4.py --alert --notify       # Alerts + notify
""")

def print_summary_card(results):
    """Print summary card for stock"""
    r = results
    upside = r['upside']
    score = r['score']

    # Rating indicator
    if upside > 30:
        rating_icon = "[BUY ]"
    elif upside > 15:
        rating_icon = "[BUY ]"
    elif upside > 0:
        rating_icon = "[HOLD]"
    else:
        rating_icon = "[SELL]"

    # Score bar
    score_bar = "=" * int(score/5) + "-" * (20 - int(score/5))

    sym = r['symbol']
    price = r['price']
    target = r['target']
    pe = r['pe']
    peg = r['peg']
    fcf = r['fcf_yield'] * 100
    roe = r['roe'] * 100
    div = r['div_yield'] * 100
    beta = r['beta']

    print("")
    print("+-------------------------------------------------------------+")
    print("| {}  ${:>8.2f}  ->  ${:>8.2f}  {} {:>+6.1f}%".format(sym, price, target, rating_icon, upside))
    print("+-------------------------------------------------------------+")
    print("| Score: [{}] {}/100".format(score_bar, score))
    print("| P/E: {:>5.1f}x  |  PEG: {:>4.2f}  |  FCF: {:>5.1f}%".format(pe, peg, fcf))
    print("| ROE: {:>5.1f}%  |  Div: {:>5.2f}%  |  Beta: {:>4.2f}".format(roe, div, beta))
    print("+-------------------------------------------------------------+")
    print("")

def interactive_mode():
    """Interactive mode"""
    print_banner()

    # Default stocks
    WATCH_LIST = ["NVDA", "META", "JPM", "AAPL", "MSFT", "GOOGL", "AMZN", "TSLA"]

    print("[ Interactive Mode - Stock PRO v10.0 ]\n")
    print("Commands:")
    print("  1-8  : Quick analyze stock")
    print("  c    : Compare all watched stocks")
    print("  w    : Watch mode (auto-refresh)")
    print("  a    : Add stock to watch")
    print("  h    : Help")
    print("  q    : Quit\n")

    print(f"Watch List: {', '.join(WATCH_LIST)}\n")

    commands = {
        "1": "NVDA", "2": "META", "3": "JPM", "4": "AAPL",
        "5": "MSFT", "6": "GOOGL", "7": "AMZN", "8": "TSLA"
    }

    while True:
        try:
            cmd = input("\n> ").strip().lower()

            if cmd == "q":
                print("\nGoodbye!")
                break
            elif cmd == "h":
                print_help()
            elif cmd == "c":
                generate_comparison(WATCH_LIST)
            elif cmd == "w":
                print("\nWatch Mode - Press Ctrl+C to exit\n")
                generate_comparison(WATCH_LIST)
            elif cmd in commands:
                sym = commands[cmd]
                print(f"\nAnalyzing {sym}...\n")
                generate_report(sym)
            elif cmd == "a":
                sym = input("Enter symbol to add: ").strip().upper()
                if sym:
                    WATCH_LIST.append(sym)
                    print(f"Added {sym} to watch list")
            else:
                # Try as symbol
                if cmd.isalpha() and len(cmd) <= 5:
                    generate_report(cmd.upper())
                else:
                    print("Unknown command. Type 'h' for help.")
        except KeyboardInterrupt:
            print("\n\nExiting watch mode...")
            break
        except Exception as e:
            print(f"Error: {e}")

def watch_mode(symbols, interval=30):
    """Watch mode - auto refresh"""
    import time

    print(f"\nWatch Mode Started - Refreshing every {interval}s")
    print("Press Ctrl+C to exit\n")

    count = 0
    while True:
        count += 1
        print(f"\n{'='*60}")
        print(f"Refresh #{count} - {datetime.now().strftime('%H:%M:%S')}")
        print('='*60)

        generate_comparison(symbols)

        print(f"\nNext refresh in {interval}s... (Ctrl+C to exit)")
        time.sleep(interval)

# ============================================================================
# INTEGRATION FUNCTIONS
# ============================================================================

def export_to_csv(symbols, output_file=None):
    """Export stock data to CSV file"""
    import csv

    if output_file is None:
        output_file = WORKSPACE / "50-reports" / "stocks" / f"export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"

    OUTPUT.mkdir(parents=True, exist_ok=True)

    headers = ["Symbol", "Price", "Target", "Upside%", "Score", "P/E", "PEG",
               "FCF%", "Div%", "ROE%", "Beta", "Rating", "Date"]

    rows = []
    for sym in symbols:
        q = fetch(sym)
        if q.get("success"):
            fin = FD.get(sym, FD["AAPL"])
            analyst = ANALYST.get(sym, {"target": q["price"]})
            dcf_scenarios = calc_dcf_scenarios(q, fin)
            target = analyst["target"] * 0.6 + dcf_scenarios["Base"] * 0.4
            upside = (target - q["price"]) / q["price"] * 100

            quality = calc_quality_score(q, fin)
            value = calc_value_score(q, fin, {"percentile": 50})
            momentum_s = calc_momentum_score(q, {"range_pos": 50, "momentum": "Neutral"})
            growth = calc_growth_score(q, fin)
            score = int(quality * 0.30 + value * 0.25 + momentum_s * 0.20 + growth * 0.25)

            rows.append([
                sym,
                f"{q['price']:.2f}",
                f"{target:.2f}",
                f"{upside:.1f}",
                score,
                f"{q['pe']:.1f}",
                f"{q['peg']:.2f}",
                f"{fin.get('fcf_yield', 0)*100:.1f}",
                f"{fin.get('div_yield', 0)*100:.2f}",
                f"{fin['roe']*100:.1f}",
                f"{q['beta']:.2f}",
                rating(upside),
                datetime.now().strftime("%Y-%m-%d")
            ])

    with open(output_file, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        writer.writerow(headers)
        writer.writerows(rows)

    print(f"\n[CSV Export] Saved to: {output_file}")
    return str(output_file)

def export_to_xlsx(symbols, output_file=None):
    """Export stock data to Excel file"""
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill
    except ImportError:
        print("\n[Excel Export] openpyxl not installed. Install with: pip install openpyxl")
        print("Falling back to CSV export...")
        return export_to_csv(symbols, output_file.replace('.xlsx', '.csv') if output_file else None)

    if output_file is None:
        OUTPUT.mkdir(parents=True, exist_ok=True)
        output_file = OUTPUT / f"export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Stock Analysis"

    # Headers
    headers = ["Symbol", "Price", "Target", "Upside%", "Score", "P/E", "PEG",
               "FCF%", "Div%", "ROE%", "Beta", "Rating", "Date"]

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Data rows
    for row_idx, sym in enumerate(symbols, 2):
        q = fetch(sym)
        if q.get("success"):
            fin = FD.get(sym, FD["AAPL"])
            analyst = ANALYST.get(sym, {"target": q["price"]})
            dcf_scenarios = calc_dcf_scenarios(q, fin)
            target = analyst["target"] * 0.6 + dcf_scenarios["Base"] * 0.4
            upside = (target - q["price"]) / q["price"] * 100

            quality = calc_quality_score(q, fin)
            value = calc_value_score(q, fin, {"percentile": 50})
            momentum_s = calc_momentum_score(q, {"range_pos": 50, "momentum": "Neutral"})
            growth = calc_growth_score(q, fin)
            score = int(quality * 0.30 + value * 0.25 + momentum_s * 0.20 + growth * 0.25)

            data = [sym, q['price'], target, upside/100, score, q['pe'], q['peg'],
                   fin.get('fcf_yield', 0), fin.get('div_yield', 0), fin['roe'],
                   q['beta'], rating(upside), datetime.now()]

            for col, val in enumerate(data, 1):
                ws.cell(row=row_idx, column=col, value=val)

    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except Exception:
                pass
        adjusted_width = min(max_length + 2, 20)
        ws.column_dimensions[column_letter].width = adjusted_width

    wb.save(output_file)
    print(f"\n[Excel Export] Saved to: {output_file}")
    return str(output_file)

def save_to_db(symbols):
    """Save stock data to SQLite database"""
    try:
        import sqlite3
    except ImportError:
        print("\n[DB] sqlite3 not available")
        return

    db_file = WORKSPACE / "50-reports" / "stocks" / "stock_data.db"
    OUTPUT.mkdir(parents=True, exist_ok=True)

    conn = sqlite3.connect(db_file)
    c = conn.cursor()

    # Create table
    c.execute('''CREATE TABLE IF NOT EXISTS stocks
                 (symbol TEXT, price REAL, target REAL, upside REAL,
                  score INTEGER, pe REAL, peg REAL, fcf_yield REAL,
                  div_yield REAL, roe REAL, beta REAL, rating TEXT,
                  timestamp TEXT)''')

    # Insert data
    for sym in symbols:
        q = fetch(sym)
        if q.get("success"):
            fin = FD.get(sym, FD["AAPL"])
            analyst = ANALYST.get(sym, {"target": q["price"]})
            dcf_scenarios = calc_dcf_scenarios(q, fin)
            target = analyst["target"] * 0.6 + dcf_scenarios["Base"] * 0.4
            upside = (target - q["price"]) / q["price"] * 100

            quality = calc_quality_score(q, fin)
            value = calc_value_score(q, fin, {"percentile": 50})
            momentum_s = calc_momentum_score(q, {"range_pos": 50, "momentum": "Neutral"})
            growth = calc_growth_score(q, fin)
            score = int(quality * 0.30 + value * 0.25 + momentum_s * 0.20 + growth * 0.25)

            c.execute('''INSERT INTO stocks VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)''',
                     (sym, q['price'], target, upside, score, q['pe'], q['peg'],
                      fin.get('fcf_yield', 0)*100, fin.get('div_yield', 0)*100,
                      fin['roe']*100, q['beta'], rating(upside),
                      datetime.now().isoformat()))

    conn.commit()
    conn.close()
    print(f"\n[Database] Saved to: {db_file}")

def generate_chart(symbols, output_file=None):
    """Generate price chart using matplotlib"""
    try:
        import matplotlib
        matplotlib.use('Agg')  # Non-interactive backend
        import matplotlib.pyplot as plt
        import matplotlib.patches as mpatches
    except ImportError:
        print("\n[Chart] matplotlib not installed. Install with: pip install matplotlib")
        return None

    if output_file is None:
        OUTPUT.mkdir(parents=True, exist_ok=True)
        output_file = OUTPUT / f"chart_{datetime.now().strftime('%Y%m%d_%H%M%S')}.png"

    # Collect data
    data = []
    for sym in symbols:
        q = fetch(sym)
        if q.get("success"):
            fin = FD.get(sym, FD["AAPL"])
            analyst = ANALYST.get(sym, {"target": q["price"]})
            dcf_scenarios = calc_dcf_scenarios(q, fin)
            target = analyst["target"] * 0.6 + dcf_scenarios["Base"] * 0.4
            upside = (target - q["price"]) / q["price"] * 100
            data.append({
                "symbol": sym,
                "price": q["price"],
                "target": target,
                "upside": upside,
                "score": 0  # Will calculate below
            })

    # Create chart
    fig, axes = plt.subplots(2, 1, figsize=(12, 8))

    # Top: Price vs Target bar chart
    symbols_list = [d["symbol"] for d in data]
    prices = [d["price"] for d in data]
    targets = [d["target"] for d in data]

    x = range(len(symbols_list))
    width = 0.35

    axes[0].bar([i - width/2 for i in x], prices, width, label='Current Price', color='#3498db')
    axes[0].bar([i + width/2 for i in x], targets, width, label='Target Price', color='#2ecc71')
    axes[0].set_ylabel('Price ($)')
    axes[0].set_title('Stock Price vs Target Price')
    axes[0].set_xticks(x)
    axes[0].set_xticklabels(symbols_list)
    axes[0].legend()
    axes[0].grid(axis='y', alpha=0.3)

    # Bottom: Upside % bar chart
    upsi = [d["upside"] for d in data]
    colors = ['#2ecc71' if u > 0 else '#e74c3c' for u in upsi]
    bars = axes[1].bar(symbols_list, upsi, color=colors)
    axes[1].set_ylabel('Upside %')
    axes[1].set_title('Potential Upside (%)')
    axes[1].axhline(y=0, color='black', linestyle='-', linewidth=0.5)
    axes[1].axhline(y=30, color='green', linestyle='--', linewidth=0.5, label='30% target')
    axes[1].grid(axis='y', alpha=0.3)

    # Add value labels
    for bar, val in zip(bars, upsi):
        height = bar.get_height()
        axes[1].annotate(f'{val:.1f}%',
                        xy=(bar.get_x() + bar.get_width() / 2, height),
                        xytext=(0, 3 if height >= 0 else -12),
                        textcoords="offset points",
                        ha='center', va='bottom' if height >= 0 else 'top',
                        fontsize=9)

    plt.tight_layout()
    plt.savefig(output_file, dpi=150, bbox_inches='tight')
    plt.close()

    print(f"\n[Chart] Saved to: {output_file}")
    return str(output_file)

def price_alert(symbols, threshold=30):
    """Check price alerts for stocks"""
    print("\n" + "="*60)
    print("Stock PRO v10.0 - Price Alert System")
    print("="*60)

    alerts = []
    for sym in symbols:
        q = fetch(sym)
        if q.get("success"):
            fin = FD.get(sym, FD["AAPL"])
            analyst = ANALYST.get(sym, {"target": q["price"]})
            dcf_scenarios = calc_dcf_scenarios(q, fin)
            target = analyst["target"] * 0.6 + dcf_scenarios["Base"] * 0.4
            upside = (target - q["price"]) / q["price"] * 100

            if upside >= threshold:
                alerts.append({
                    "symbol": sym,
                    "price": q["price"],
                    "target": target,
                    "upside": upside
                })

    if alerts:
        print(f"\n[!] {len(alerts)} stocks with upside >= {threshold}%:\n")
        print(f"{'Symbol':<8} {'Price':>10} {'Target':>10} {'Upside':>10}")
        print("-" * 40)
        for a in alerts:
            print(f"{a['symbol']:<8} ${a['price']:>9.2f} ${a['target']:>9.2f} {a['upside']:>+9.1f}%")

        print("\n[ALERT] Consider buying these stocks!")
        return alerts
    else:
        print(f"\n[-] No stocks with upside >= {threshold}%")
        return []

def sync_to_obsidian(symbols):
    """Sync stock analysis to Obsidian vault"""
    obsidian_path = WORKSPACE / "50-reports" / "obsidian" / "stocks"
    obsidian_path.mkdir(parents=True, exist_ok=True)

    for sym in symbols:
        q = fetch(sym)
        if q.get("success"):
            fin = FD.get(sym, FD["AAPL"])
            analyst = ANALYST.get(sym, {"target": q["price"]})
            dcf_scenarios = calc_dcf_scenarios(q, fin)
            target = analyst["target"] * 0.6 + dcf_scenarios["Base"] * 0.4
            upside = (target - q["price"]) / q["price"] * 100

            quality = calc_quality_score(q, fin)
            value = calc_value_score(q, fin, {"percentile": 50})
            momentum_s = calc_momentum_score(q, {"range_pos": 50, "momentum": "Neutral"})
            growth = calc_growth_score(q, fin)
            score = int(quality * 0.30 + value * 0.25 + momentum_s * 0.20 + growth * 0.25)

            price_val = q['price']
            target_val = target
            upside_val = upside
            pe_val = q['pe']
            peg_val = q['peg']
            fpe_val = q['pe']/1.5
            dcf_bull_val = dcf_scenarios['Bull']
            dcf_base_val = dcf_scenarios['Base']
            dcf_bear_val = dcf_scenarios['Bear']
            roe_val = fin['roe']*100
            pm_val = fin.get('profit_margin', 0.25)*100
            de_val = fin.get('debt_equity', 0.5)
            fcf_val = fin.get('fcf_yield', 0)*100
            div_val = fin.get('div_yield', 0)*100
            beta_val = q['beta']
            high52 = q['price']*1.3
            low52 = q['price']*0.7
            rating_val = rating(upside)

            # Build markdown content using concatenation
            md_content = "---\n"
            md_content += "type: stock-analysis\n"
            md_content += "symbol: " + sym + "\n"
            md_content += "date: " + datetime.now().strftime('%Y-%m-%d') + "\n"
            md_content += "tags: [stocks, analysis]\n"
            md_content += "---\n\n"
            md_content += "# " + sym + " Stock Analysis\n\n"
            md_content += "## Quick Stats\n"
            md_content += "| Metric | Value |\n"
            md_content += "|--------|-------|\n"
            md_content += "| Price | ${:.2f} |\n".format(price_val)
            md_content += "| Target | ${:.2f} |\n".format(target_val)
            md_content += "| Upside | {:+.1f}% |\n".format(upside_val)
            md_content += "| Rating | {} |\n".format(rating_val)
            md_content += "| Score | {}/100 |\n\n".format(score)
            md_content += "## Valuation\n"
            md_content += "| Metric | Value |\n"
            md_content += "|--------|-------|\n"
            md_content += "| P/E | {:.1f}x |\n".format(pe_val)
            md_content += "| Forward P/E | {:.1f}x |\n".format(fpe_val)
            md_content += "| PEG | {:.2f} |\n".format(peg_val)
            md_content += "| DCF Bull | ${:.2f} |\n".format(dcf_bull_val)
            md_content += "| DCF Base | ${:.2f} |\n".format(dcf_base_val)
            md_content += "| DCF Bear | ${:.2f} |\n\n".format(dcf_bear_val)
            md_content += "## Fundamentals\n"
            md_content += "| Metric | Value |\n"
            md_content += "|--------|-------|\n"
            md_content += "| ROE | {:.1f}% |\n".format(roe_val)
            md_content += "| Profit Margin | {:.1f}% |\n".format(pm_val)
            md_content += "| Debt/Equity | {:.2f} |\n".format(de_val)
            md_content += "| FCF Yield | {:.1f}% |\n".format(fcf_val)
            md_content += "| Div Yield | {:.2f}% |\n\n".format(div_val)
            md_content += "## Technical\n"
            md_content += "| Metric | Value |\n"
            md_content += "|--------|-------|\n"
            md_content += "| Beta | {:.2f} |\n".format(beta_val)
            md_content += "| 52W High | ${:.2f} |\n".format(high52)
            md_content += "| 52W Low | ${:.2f} |\n".format(low52)

            # Add analyst consensus section
            md_content += "\n## Analyst Consensus\n"
            md_content += "- Target: ${:.2f}\n".format(analyst['target'])
            md_content += "- Rating: {}\n".format(analyst['rating'])
            md_content += "- Number of Analysts: {}\n\n".format(analyst.get('num', 'N/A'))
            md_content += "## Investment Thesis\n\n"
            md_content += "### Strengths\n- \n\n"
            md_content += "### Weaknesses\n- \n\n"
            md_content += "### Opportunities\n- \n\n"
            md_content += "### Threats\n- \n\n"
            md_content += "---\n"
            md_content += "*Generated by Stock PRO v10.0 on "
            md_content += datetime.now().strftime('%Y-%m-%d %H:%M:%S') + "*\n"

            # Save markdown file
            md_file = obsidian_path / f"{sym}.md"
            with open(md_file, 'w', encoding='utf-8') as f:
                f.write(md_content)

    print(f"\n[Obsidian] Synced {len(symbols)} stocks to: {obsidian_path}")
    return str(obsidian_path)

def setup_cron(cron_expr, symbols):
    """Setup cron job for automatic stock analysis"""
    import os

    cron_file = WORKSPACE / "30-scripts-tools" / "stock_cron.bat"

    # Create Windows Task Scheduler command
    symbols_str = " ".join(symbols)
    cmd = f'py "{WORKSPACE}\\30-scripts-tools\\stock_pro_v4.py" --compare {symbols_str}'

    # Read existing cron skill
    cron_skill = WORKSPACE / "active_skills" / "cron" / "SKILL.md"
    if cron_skill.exists():
        with open(cron_skill, 'r', encoding='utf-8') as f:
            content = f.read()

        # Parse cron expression
        parts = cron_expr.split()
        if len(parts) >= 5:
            minute, hour, day, month, dow = parts[:5]
            desc = f"Stock analysis at {hour}:{minute.zfill(2)} on {dow}"

            # Add task entry
            task_entry = f"""
### Stock Analysis - {datetime.now().strftime('%Y-%m-%d')}
- Command: `{cmd}`
- Schedule: {cron_expr}
- Description: {desc}
- Status: PENDING
"""

            # Save to cron log
            log_file = WORKSPACE / "30-scripts-tools" / "cron_log.md"
            with open(log_file, 'a', encoding='utf-8') as f:
                f.write(task_entry)

            print(f"\n[Cron] Task scheduled:")
            print(f"  Command: {cmd}")
            print(f"  Schedule: {cron_expr}")
            print(f"  Log: {log_file}")
            return True
    else:
        print(f"\n[Cron] Cron skill not found. Using Task Scheduler...")

        # Windows Task Scheduler alternative
        task_name = "StockPRO_AutoAnalysis"
        schedule_cmd = f'schtasks /create /tn "{task_name}" /tr "{cmd}" /sc daily /st {hour}:{minute}'

        with open(cron_file, 'w', encoding='utf-8') as f:
            f.write(f'@echo off\n{schedule_cmd}\npause')

        print(f"\n[Cron] Created scheduler script: {cron_file}")
        print("Run as Administrator to create the scheduled task.")
        return True

# ============================================================================
# ADVANCED INTEGRATIONS
# ============================================================================

def start_api_server(port=8765):
    """Start a simple API server for stock data"""
    try:
        from http.server import HTTPServer, BaseHTTPRequestHandler
        import urllib.parse
    except ImportError:
        print("\n[API] Failed to start server")
        return

    config = load_config()

    class StockHandler(BaseHTTPRequestHandler):
        def do_GET(self):
            parsed = urllib.parse.urlparse(self.path)
            path = parsed.path

            if path == '/health':
                self.send_response(200)
                self.send_header('Content-type', 'application/json')
                self.end_headers()
                self.wfile.write(json.dumps({"status": "ok", "time": datetime.now().isoformat()}).encode())

            elif path == '/stocks' or path == '/':
                symbols = config.get('watchlist', ['NVDA', 'META', 'JPM'])
                results = []
                for sym in symbols:
                    q = fetch(sym)
                    if q.get("success"):
                        fin = FD.get(sym, FD["AAPL"])
                        analyst = ANALYST.get(sym, {"target": q["price"]})
                        dcf_scenarios = calc_dcf_scenarios(q, fin)
                        target = analyst["target"] * 0.6 + dcf_scenarios["Base"] * 0.4
                        upside = (target - q["price"]) / q["price"] * 100
                        results.append({
                            "symbol": sym,
                            "price": q["price"],
                            "target": round(target, 2),
                            "upside": round(upside, 1),
                            "rating": rating(upside)
                        })

                self.send_response(200)
                self.send_header('Content-type', 'application/json')
                self.end_headers()
                self.wfile.write(json.dumps(results, indent=2).encode())

            elif path.startswith('/stock/'):
                sym = path.split('/')[-1].upper()
                q = fetch(sym)
                if q.get("success"):
                    fin = FD.get(sym, FD["AAPL"])
                    analyst = ANALYST.get(sym, {"target": q["price"]})
                    dcf_scenarios = calc_dcf_scenarios(q, fin)
                    target = analyst["target"] * 0.6 + dcf_scenarios["Base"] * 0.4
                    upside = (target - q["price"]) / q["price"] * 100

                    quality = calc_quality_score(q, fin)
                    value = calc_value_score(q, fin, {"percentile": 50})
                    momentum_s = calc_momentum_score(q, {"range_pos": 50, "momentum": "Neutral"})
                    growth = calc_growth_score(q, fin)
                    score = int(quality * 0.30 + value * 0.25 + momentum_s * 0.20 + growth * 0.25)

                    result = {
                        "symbol": sym,
                        "price": q["price"],
                        "target": round(target, 2),
                        "upside": round(upside, 1),
                        "score": score,
                        "rating": rating(upside),
                        "metrics": {
                            "pe": q["pe"],
                            "peg": q["peg"],
                            "fcf_yield": fin.get("fcf_yield", 0),
                            "div_yield": fin.get("div_yield", 0),
                            "roe": fin["roe"],
                            "beta": q["beta"],
                        },
                        "dcf": {k: round(v, 2) for k, v in dcf_scenarios.items()},
                        "analyst": analyst
                    }

                    self.send_response(200)
                    self.send_header('Content-type', 'application/json')
                    self.end_headers()
                    self.wfile.write(json.dumps(result, indent=2).encode())
                else:
                    self.send_response(404)
                    self.send_header('Content-type', 'application/json')
                    self.end_headers()
                    self.wfile.write(json.dumps({"error": "Stock not found"}).encode())

            else:
                self.send_response(404)
                self.end_headers()

        def log_message(self, format, *args):
            print(f"[API] {args[0]}")

    print(f"\n[API] Starting server on port {port}...")
    print(f"[API] Endpoints:")
    print(f"  GET /              - List all stocks")
    print(f"  GET /stocks        - List all stocks")
    print(f"  GET /stock/NVDA    - Get NVDA details")
    print(f"  GET /health        - Health check")
    print(f"\n[API] Press Ctrl+C to stop\n")

    try:
        server = HTTPServer(('localhost', port), StockHandler)
        server.serve_forever()
    except KeyboardInterrupt:
        print("\n[API] Server stopped")
        server.shutdown()

def send_webhook(url, data):
    """Send data to webhook URL"""
    try:
        import urllib.request

        payload = json.dumps({
            "event": "stock_alert",
            "timestamp": datetime.now().isoformat(),
            "data": data
        }).encode('utf-8')

        req = urllib.request.Request(
            url,
            data=payload,
            headers={'Content-Type': 'application/json'}
        )

        with urllib.request.urlopen(req, timeout=10) as response:
            print(f"[Webhook] Sent successfully: {response.status}")
            return True
    except Exception as e:
        print(f"[Webhook] Failed: {e}")
        return False

def notify_alerts(symbols, threshold=30):
    """Send alert notifications"""
    config = load_config()
    alerts = price_alert(symbols, threshold)

    if not alerts:
        return

    # Send webhook if configured
    if config.get('notifications', {}).get('enabled'):
        webhook_url = config.get('notifications', {}).get('webhook_url', '')
        if webhook_url:
            send_webhook(webhook_url, {"alerts": alerts, "count": len(alerts)})

    # Save alert to file
    alert_file = OUTPUT / f"alerts_{datetime.now().strftime('%Y%m%d')}.json"
    with open(alert_file, 'w', encoding='utf-8') as f:
        json.dump({"alerts": alerts, "timestamp": datetime.now().isoformat()}, f, indent=2)

    print(f"[Alert] Saved to: {alert_file}")

def create_pipeline_script(symbols, output_file=None):
    """Create a batch pipeline script"""
    if output_file is None:
        output_file = WORKSPACE / "30-scripts-tools" / "stock_pipeline.bat"

    symbols_str = " ".join(symbols)

    script = f'''@echo off
REM Stock PRO Pipeline - {datetime.now().strftime('%Y-%m-%d %H:%M')}
REM Auto-generated by Stock PRO v10.0

echo.
echo ============================================================
echo Stock PRO Pipeline - {symbols_str}
echo ============================================================
echo.

REM 1. Generate comparison report
echo [1/4] Generating comparison report...
python "%~dp0stock_pro_v4.py" --compare {symbols_str}

REM 2. Export to CSV
echo.
echo [2/4] Exporting to CSV...
python "%~dp0stock_pro_v4.py" --csv {symbols_str}

REM 3. Check alerts
echo.
echo [3/4] Checking price alerts...
python "%~dp0stock_pro_v4.py" --alert {symbols_str}

REM 4. Sync to Obsidian
echo.
echo [4/4] Syncing to Obsidian...
python "%~dp0stock_pro_v4.py" --obsidian {symbols_str}

echo.
echo ============================================================
echo Pipeline Complete!
echo ============================================================
pause
'''

    with open(output_file, 'w', encoding='utf-8') as f:
        f.write(script)

    print(f"\n[Pipeline] Created: {output_file}")
    print("Run with: stock_pipeline.bat")
    return str(output_file)

def create_copaw_skill():
    """Create CoPaw skill for stock analysis"""
    skill_dir = WORKSPACE / "active_skills" / "stock-pro"
    skill_dir.mkdir(parents=True, exist_ok=True)

    # SKILL.md - using list to build content
    skill_md_lines = [
        "# Stock PRO Skill",
        "",
        "## Description",
        "Professional stock analysis tool with valuation, DCF, technical analysis, and multi-stock comparison.",
        "",
        "## Usage",
        "```",
        "stock-pro <symbol> [--cn] [--json]",
        "stock-pro --compare <symbols...>",
        "stock-pro --watch <symbols...>",
        "stock-pro --alert <symbols...>",
        "stock-pro --chart <symbols...>",
        "stock-pro --csv <symbols...>",
        "stock-pro --xlsx <symbols...>",
        "stock-pro --db <symbols...>",
        "stock-pro --obsidian <symbols...>",
        "```",
        "",
        "## Options",
        "- `--cn` Chinese report",
        "- `--json` JSON output",
        "- `--compare` Multi-stock comparison",
        "- `--watch` Auto-refresh mode",
        "- `--alert` Price alerts (upside > 30%)",
        "- `--chart` Generate chart",
        "- `--csv` Export to CSV",
        "- `--xlsx` Export to Excel",
        "- `--db` Save to SQLite",
        "- `--obsidian` Sync to Obsidian",
        "",
        "## Examples",
        "```bash",
        "stock-pro NVDA              # English report",
        "stock-pro NVDA --cn         # Chinese report",
        "stock-pro --compare NVDA META JPM",
        "stock-pro --alert NVDA META JPM",
        "stock-pro --watch NVDA META",
        "```",
        "",
        "## Output",
        "- Reports: 50-reports/stocks/*.md",
        "- Charts: 50-reports/stocks/*.png",
        "- Data: 50-reports/stocks/*.csv, *.xlsx, *.db",
        "",
        "## CoPaw Integration",
        "This skill can be triggered by:",
        '- "analyze NVDA"',
        '- "stock comparison NVDA META"',
        '- "watch NVDA META"',
        "",
    ]
    skill_md = "\n".join(skill_md_lines)

    # skill.sh / skill.bat
    skill_sh_lines = [
        "#!/bin/bash",
        "# Stock PRO Skill for CoPaw",
        'SCRIPT_DIR="$(cd "$(dirname "${BASH_SOURCE[0]}")" && pwd)"',
        'python "$SCRIPT_DIR/../30-scripts-tools/stock_pro_v4.py" "$@"',
        "",
    ]
    skill_sh = "\n".join(skill_sh_lines)

    skill_bat = '@echo off\r\npython "%~dp0..\\30-scripts-tools\\stock_pro_v4.py" %*\r\n'

    # Write files
    with open(skill_dir / "SKILL.md", 'w', encoding='utf-8') as f:
        f.write(skill_md)

    with open(skill_dir / "skill.sh", 'w', encoding='utf-8') as f:
        f.write(skill_sh)

    with open(skill_dir / "skill.bat", 'w', encoding='utf-8') as f:
        f.write(skill_bat)

    print(f"\n[CoPaw Skill] Created at: {skill_dir}")
    print("Enable by adding 'stock-pro' to your active skills")
    return str(skill_dir)

# ============================================================================
# PORTFOLIO & NEWS MODULES
# ============================================================================

def load_portfolio():
    """Load portfolio from file"""
    portfolio_file = OUTPUT / "portfolio.json"
    if portfolio_file.exists():
        try:
            with open(portfolio_file, 'r', encoding='utf-8') as f:
                return json.load(f)
        except Exception:
            pass
    return {"positions": [], "cash": 10000, "currency": "USD"}

def save_portfolio(portfolio):
    """Save portfolio to file"""
    portfolio_file = OUTPUT / "portfolio.json"
    with open(portfolio_file, 'w', encoding='utf-8') as f:
        json.dump(portfolio, f, indent=2)

def add_position(symbol, shares, avg_cost):
    """Add a position to portfolio"""
    portfolio = load_portfolio()
    portfolio["positions"].append({
        "symbol": symbol.upper(),
        "shares": float(shares),
        "avg_cost": float(avg_cost),
        "date_added": datetime.now().isoformat()
    })
    save_portfolio(portfolio)
    print(f"\n[Portfolio] Added {shares} shares of {symbol} @ ${avg_cost}")

def remove_position(symbol):
    """Remove a position from portfolio"""
    portfolio = load_portfolio()
    portfolio["positions"] = [p for p in portfolio["positions"] if p["symbol"] != symbol.upper()]
    save_portfolio(portfolio)
    print(f"\n[Portfolio] Removed {symbol}")

def show_portfolio():
    """Show portfolio with current values"""
    portfolio = load_portfolio()

    if not portfolio["positions"]:
        print("\n[Portfolio] Empty. Add positions with: --portfolio-add NVDA 100 170")
        return

    print("\n" + "="*80)
    print("Stock PRO Portfolio")
    print("="*80)
    print(f"{'Symbol':<8} {'Shares':>8} {'Avg Cost':>10} {'Current':>10} {'Value':>12} {'Gain/Loss':>12}")
    print("-"*80)

    total_value = portfolio.get("cash", 0)
    total_cost = 0

    for pos in portfolio["positions"]:
        sym = pos["symbol"]
        shares = pos["shares"]
        avg_cost = pos["avg_cost"]

        q = fetch(sym)
        current = q["price"] if q.get("success") else avg_cost

        value = shares * current
        cost = shares * avg_cost
        gain_loss = value - cost
        gain_pct = (gain_loss / cost * 100) if cost > 0 else 0

        total_value += value
        total_cost += cost

        color = "\033[92m" if gain_loss >= 0 else "\033[91m"
        reset = "\033[0m"
        sign = "+" if gain_loss >= 0 else ""

        print(f"{sym:<8} {shares:>8.1f} ${avg_cost:>9.2f} ${current:>9.2f} ${value:>11.2f} {color}{sign}{gain_loss:>10.2f} ({gain_pct:>+5.1f}%){reset}")

    print("-"*80)
    total_gain_loss = total_value - total_cost
    total_gain_pct = (total_gain_loss / total_cost * 100) if total_cost > 0 else 0
    color = "\033[92m" if total_gain_loss >= 0 else "\033[91m"
    reset = "\033[0m"
    sign = "+" if total_gain_loss >= 0 else ""

    print(f"{'Cash':<8} ${portfolio.get('cash', 0):>67.2f}")
    print(f"{'TOTAL':<8} {'':<8} {'':<10} {'':<10} ${total_value:>11.2f} {color}{sign}{total_gain_loss:>10.2f} ({total_gain_pct:>+5.1f}%){reset}")
    print("="*80)

def fetch_news(symbols):
    """Fetch news for stocks from cache or API"""
    cache_file = NEWS_CACHE_FILE
    news_data = {"news": [], "timestamp": None}

    # Load cache
    if cache_file.exists():
        try:
            with open(cache_file, 'r', encoding='utf-8') as f:
                news_data = json.load(f)

            # Check if cache is fresh (within 1 hour)
            if news_data.get("timestamp"):
                cache_age = datetime.now() - datetime.fromisoformat(news_data["timestamp"])
                if cache_age < timedelta(hours=1):
                    return news_data["news"]
        except Exception:
            pass

    # Fetch fresh news (mock for now - in production would use news API)
    news = []
    for sym in symbols:
        news.append({
            "symbol": sym,
            "title": f"{sym} Stock News",
            "source": "Market Data",
            "url": f"https://finance.yahoo.com/quote/{sym}",
            "published": datetime.now().isoformat(),
            "sentiment": "neutral"
        })

    # Save cache
    with open(cache_file, 'w', encoding='utf-8') as f:
        json.dump({"news": news, "timestamp": datetime.now().isoformat()}, f)

    return news

def show_news(symbols):
    """Show news for stocks"""
    news = fetch_news(symbols)

    print("\n" + "="*80)
    print("Stock PRO News Feed")
    print("="*80 + "\n")

    for item in news:
        sentiment_emoji = "🟢" if item["sentiment"] == "positive" else ("🔴" if item["sentiment"] == "negative" else "⚪")
        print(f"{sentiment_emoji} {item['symbol']}: {item['title']}")
        print(f"   Source: {item['source']} | {item['published'][:10]}")
        print()

def screen_stocks(criteria=None):
    """Screen stocks based on criteria"""
    if criteria is None:
        criteria = {
            "min_score": 60,
            "max_pe": 40,
            "min_upside": 15,
            "min_roe": 10
        }

    # Get all known stocks
    all_symbols = list(ANALYST.keys())
    results = []

    for sym in all_symbols:
        q = fetch(sym)
        if not q.get("success"):
            continue

        fin = FD.get(sym, FD.get("AAPL", {}))
        analyst = ANALYST.get(sym, {"target": q["price"]})
        dcf_scenarios = calc_dcf_scenarios(q, fin)
        target = analyst["target"] * 0.6 + dcf_scenarios["Base"] * 0.4
        upside = (target - q["price"]) / q["price"] * 100

        quality = calc_quality_score(q, fin)
        value = calc_value_score(q, fin, {"percentile": 50})
        momentum_s = calc_momentum_score(q, {"range_pos": 50, "momentum": "Neutral"})
        growth = calc_growth_score(q, fin)
        score = int(quality * 0.30 + value * 0.25 + momentum_s * 0.20 + growth * 0.25)

        # Apply filters
        if score < criteria.get("min_score", 0):
            continue
        if q["pe"] > criteria.get("max_pe", 999):
            continue
        if upside < criteria.get("min_upside", -999):
            continue
        if fin.get("roe", 0) * 100 < criteria.get("min_roe", 0):
            continue

        results.append({
            "symbol": sym,
            "score": score,
            "upside": upside,
            "pe": q["pe"],
            "peg": q["peg"],
            "roe": fin.get("roe", 0) * 100,
            "fcf": fin.get("fcf_yield", 0) * 100,
            "div": fin.get("div_yield", 0) * 100,
        })

    # Sort by score
    results.sort(key=lambda x: x["score"], reverse=True)

    return results

def show_screener(min_score=60, min_upside=15, max_pe=40):
    """Show stock screener results"""
    criteria = {
        "min_score": min_score,
        "min_upside": min_upside,
        "max_pe": max_pe,
        "min_roe": 10
    }

    results = screen_stocks(criteria)

    print("\n" + "="*100)
    print(f"Stock Screener (Score>={min_score}, Upside>={min_upside}%, P/E<={max_pe})")
    print("="*100)
    print(f"{'Symbol':<8} {'Score':>6} {'Upside':>8} {'P/E':>6} {'PEG':>6} {'ROE':>6} {'FCF':>6} {'Div':>6}")
    print("-"*100)

    if not results:
        print("No stocks match criteria. Try relaxing filters.")
        return

    for r in results:
        print(f"{r['symbol']:<8} {r['score']:>6} {r['upside']:>+7.1f}% {r['pe']:>5.1f}x {r['peg']:>5.2f} {r['roe']:>5.1f}% {r['fcf']:>5.1f}% {r['div']:>5.2f}%")

    print("-"*100)
    print(f"Found {len(results)} stocks matching criteria")

def send_email_report(symbols, recipients=None):
    """Send stock report via email"""
    config = load_config()
    email_config = config.get("email", {})

    if not email_config.get("enabled") and not recipients:
        print("\n[Email] Email not configured. Set up in config:")
        print("  stock_pro_v4.py --config")
        print("  Then add email settings to stock_pro_config.json")
        return

    # Generate report content
    data = generate_dashboard_data(symbols)

    # Build email body
    body = "Stock PRO Daily Report\n"
    body += "="*60 + "\n\n"

    for d in data:
        body += f"{d['symbol']}: ${d['price']:.2f} -> ${d['target']:.2f} ({d['upside']:+.1f}%)\n"
        body += f"  Rating: {d['rating']} | Score: {d['score']}/100\n"
        body += f"  P/E: {d['pe']:.1f}x | ROE: {d['roe']:.1f}% | FCF: {d['fcf']:.1f}%\n\n"

    body += f"\nGenerated: {datetime.now().strftime('%Y-%m-%d %H:%M')}\n"
    body += "Stock PRO v10.0\n"

    # Try to send email
    try:
        import smtplib
        from email.mime.text import MIMEText
        from email.mime.multipart import MIMEMultipart

        smtp_server = email_config.get("smtp_server", "smtp.gmail.com")
        smtp_port = email_config.get("smtp_port", 587)
        smtp_user = email_config.get("smtp_user", "")
        smtp_password = email_config.get("smtp_password", "")

        if not smtp_user or not recipients:
            print("\n[Email] SMTP not configured. Install python-dotenv and configure email settings.")
            print("Email content generated (not sent):")
            print(body[:500] + "...")
            return

        msg = MIMEMultipart()
        msg['From'] = smtp_user
        msg['To'] = ", ".join(recipients)
        msg['Subject'] = f"Stock PRO Report - {datetime.now().strftime('%Y-%m-%d')}"

        msg.attach(MIMEText(body, 'plain'))

        with smtplib.SMTP(smtp_server, smtp_port) as server:
            server.starttls()
            server.login(smtp_user, smtp_password)
            server.send_message(msg)

        print(f"\n[Email] Report sent to: {', '.join(recipients)}")
    except Exception as e:
        print(f"\n[Email] Failed to send: {e}")
        print("Email content generated (not sent):")
        print(body[:500] + "...")

def generate_dashboard_data(symbols):
    data = []
    for sym in symbols:
        q = fetch(sym)
        if q.get("success"):
            fin = FD.get(sym, FD["AAPL"])
            analyst = ANALYST.get(sym, {"target": q["price"]})
            dcf_scenarios = calc_dcf_scenarios(q, fin)
            target = analyst["target"] * 0.6 + dcf_scenarios["Base"] * 0.4
            upside = (target - q["price"]) / q["price"] * 100

            quality = calc_quality_score(q, fin)
            value = calc_value_score(q, fin, {"percentile": 50})
            momentum_s = calc_momentum_score(q, {"range_pos": 50, "momentum": "Neutral"})
            growth = calc_growth_score(q, fin)
            score = int(quality * 0.30 + value * 0.25 + momentum_s * 0.20 + growth * 0.25)

            data.append({
                "symbol": sym,
                "price": q["price"],
                "target": round(target, 2),
                "upside": round(upside, 1),
                "score": score,
                "rating": rating(upside),
                "pe": q["pe"],
                "peg": q["peg"],
                "roe": round(fin["roe"] * 100, 1),
                "fcf": round(fin.get("fcf_yield", 0) * 100, 1),
                "div": round(fin.get("div_yield", 0) * 100, 2),
                "beta": q["beta"],
                "dcf_bull": round(dcf_scenarios["Bull"], 2),
                "dcf_base": round(dcf_scenarios["Base"], 2),
                "dcf_bear": round(dcf_scenarios["Bear"], 2),
                "analyst_target": analyst["target"],
                "analyst_rating": analyst["rating"],
                "timestamp": datetime.now().isoformat()
            })

    # Sort by upside
    data.sort(key=lambda x: x["upside"], reverse=True)

    return data

def create_dashboard_html(symbols):
    """Create HTML dashboard"""
    data = generate_dashboard_data(symbols)

    # Calculate summary stats
    top_pick = data[0] if data else None
    avg_upside = sum(d["upside"] for d in data) / len(data) if data else 0
    strong_buys = len([d for d in data if d["upside"] > 30])

    html = f'''<!DOCTYPE html>
<html>
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Stock PRO Dashboard</title>
    <style>
        * {{ margin: 0; padding: 0; box-sizing: border-box; }}
        body {{ font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif; background: #0a0a0a; color: #fff; padding: 20px; }}
        .header {{ display: flex; justify-content: space-between; align-items: center; margin-bottom: 30px; }}
        .header h1 {{ color: #00d4ff; }}
        .header .time {{ color: #888; }}
        .stats {{ display: grid; grid-template-columns: repeat(auto-fit, minmax(200px, 1fr)); gap: 20px; margin-bottom: 30px; }}
        .stat-card {{ background: #1a1a2e; padding: 20px; border-radius: 10px; }}
        .stat-card .label {{ color: #888; font-size: 14px; }}
        .stat-card .value {{ font-size: 32px; font-weight: bold; color: #00d4ff; }}
        .stat-card.top .value {{ color: #00ff88; }}
        table {{ width: 100%; border-collapse: collapse; background: #1a1a2e; border-radius: 10px; overflow: hidden; }}
        th {{ background: #2d2d4a; padding: 15px; text-align: left; color: #888; }}
        td {{ padding: 15px; border-bottom: 1px solid #2d2d4a; }}
        tr:hover {{ background: #252540; }}
        .buy {{ color: #00ff88; }}
        .hold {{ color: #ffcc00; }}
        .sell {{ color: #ff4444; }}
        .positive {{ color: #00ff88; }}
        .negative {{ color: #ff4444; }}
        .refresh {{ background: #00d4ff; color: #000; padding: 10px 20px; border: none; border-radius: 5px; cursor: pointer; }}
        .refresh:hover {{ background: #00ffcc; }}
    </style>
</head>
<body>
    <div class="header">
        <h1>Stock PRO Dashboard</h1>
        <div>
            <span class="time">Updated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</span>
            <button class="refresh" onclick="location.reload()">Refresh</button>
        </div>
    </div>
    
    <div class="stats">
        <div class="stat-card top">
            <div class="label">Top Pick</div>
            <div class="value">{top_pick['symbol'] if top_pick else 'N/A'}</div>
        </div>
        <div class="stat-card">
            <div class="label">Avg Upside</div>
            <div class="value">{avg_upside:.1f}%</div>
        </div>
        <div class="stat-card">
            <div class="label">Strong Buys</div>
            <div class="value">{strong_buys}</div>
        </div>
        <div class="stat-card">
            <div class="label">Stocks</div>
            <div class="value">{len(data)}</div>
        </div>
    </div>
    
    <table>
        <thead>
            <tr>
                <th>Symbol</th>
                <th>Price</th>
                <th>Target</th>
                <th>Upside</th>
                <th>Score</th>
                <th>Rating</th>
                <th>P/E</th>
                <th>ROE</th>
                <th>FCF</th>
            </tr>
        </thead>
        <tbody>
'''

    for d in data:
        rating_class = "buy" if d["upside"] > 30 else ("hold" if d["upside"] > 0 else "sell")
        upside_class = "positive" if d["upside"] > 0 else "negative"
        html += f'''            <tr>
                <td><strong>{d["symbol"]}</strong></td>
                <td>${d["price"]:.2f}</td>
                <td>${d["target"]:.2f}</td>
                <td class="{upside_class}">{d["upside"]:+.1f}%</td>
                <td>{d["score"]}/100</td>
                <td class="{rating_class}">{d["rating"]}</td>
                <td>{d["pe"]:.1f}x</td>
                <td>{d["roe"]:.1f}%</td>
                <td>{d["fcf"]:.1f}%</td>
            </tr>
'''

    html += '''        </tbody>
    </table>
    
    <script>
        // Auto-refresh every 5 minutes
        setTimeout(() => location.reload(), 300000);
    </script>
</body>
</html>
'''

    # Save dashboard
    dashboard_file = OUTPUT / "dashboard.html"
    with open(dashboard_file, 'w', encoding='utf-8') as f:
        f.write(html)

    print(f"\n[Dashboard] Created: {dashboard_file}")
    return str(dashboard_file)

def generate_json_report(sym):
    """Generate JSON format report"""
    q = fetch(sym)
    if not q.get("success"):
        return {"error": "Failed to fetch data"}

    fin = FD.get(sym, FD["AAPL"])
    analyst = ANALYST.get(sym, {"target": q["price"], "rating": "N/A"})

    price = q["price"]
    analyst_target = analyst["target"]
    dcf_scenarios = calc_dcf_scenarios(q, fin)
    dcf_base = dcf_scenarios["Base"]
    target = analyst_target * 0.6 + dcf_base * 0.4
    upside = (target - price) / price * 100

    quality = calc_quality_score(q, fin)
    value = calc_value_score(q, fin, {"percentile": 50})
    momentum_s = calc_momentum_score(q, {"range_pos": 50, "momentum": "Neutral"})
    growth = calc_growth_score(q, fin)
    sc = int(quality * 0.30 + value * 0.25 + momentum_s * 0.20 + growth * 0.25)

    return {
        "symbol": sym,
        "price": price,
        "target": target,
        "upside": upside,
        "score": sc,
        "rating": rating(upside),
        "metrics": {
            "pe": q["pe"],
            "peg": q["peg"],
            "fcf_yield": fin.get("fcf_yield", 0),
            "div_yield": fin.get("div_yield", 0),
            "roe": fin["roe"],
            "beta": q["beta"],
        },
        "dcf": dcf_scenarios,
        "analyst_target": analyst_target,
        "timestamp": datetime.now().isoformat(),
    }

if __name__ == "__main__":
    import sys

    # Show banner by default
    print_banner()

    # Check for help
    if "--help" in sys.argv or "-h" in sys.argv:
        print_help()
        sys.exit(0)

    # Get symbols (uppercase, no flags)
    symbols = [arg.upper() for arg in sys.argv[1:] if arg.isalpha() and len(arg) <= 10]
    if not symbols:
        symbols = ["NVDA", "META", "JPM", "AAPL", "MSFT", "GOOGL"]

    # Check for interactive mode
    if "--interactive" in sys.argv or "-i" in sys.argv:
        interactive_mode()
        sys.exit(0)

    # Check for watch mode
    if "--watch" in sys.argv or "-w" in sys.argv:
        watch_mode(symbols)
        sys.exit(0)

    # Check for alert mode
    if "--alert" in sys.argv or "-a" in sys.argv:
        threshold = 30
        if "--threshold" in sys.argv:
            idx = sys.argv.index("--threshold") + 1
            if idx < len(sys.argv):
                threshold = float(sys.argv[idx])
        price_alert(symbols, threshold)
        sys.exit(0)

    # Check for JSON mode
    if "--json" in sys.argv or "-j" in sys.argv:
        import json
        result = generate_json_report(symbols[0])
        print(json.dumps(result, indent=2))
        sys.exit(0)

    # Check for comparison mode
    if "--compare" in sys.argv or "-c" in sys.argv:
        generate_comparison(symbols)
        sys.exit(0)

    # Check for CSV export
    if "--csv" in sys.argv:
        export_to_csv(symbols)
        sys.exit(0)

    # Check for Excel export
    if "--xlsx" in sys.argv or "--excel" in sys.argv:
        export_to_xlsx(symbols)
        sys.exit(0)

    # Check for database save
    if "--db" in sys.argv:
        save_to_db(symbols)
        sys.exit(0)

    # Check for chart generation
    if "--chart" in sys.argv:
        generate_chart(symbols)
        sys.exit(0)

    # Check for Obsidian sync
    if "--obsidian" in sys.argv:
        sync_to_obsidian(symbols)
        sys.exit(0)

    # Check for cron setup
    if "--cron" in sys.argv:
        cron_expr = "0 9 * * 1-5"  # Default: 9AM on weekdays
        if "--cron" in sys.argv:
            idx = sys.argv.index("--cron") + 1
            if idx < len(sys.argv) and not sys.argv[idx].startswith("-"):
                cron_expr = sys.argv[idx]
        setup_cron(cron_expr, symbols)
        sys.exit(0)

    # Check for API server mode
    if "--api" in sys.argv or "--server" in sys.argv:
        port = 8765
        if "--port" in sys.argv:
            idx = sys.argv.index("--port") + 1
            if idx < len(sys.argv):
                port = int(sys.argv[idx])
        start_api_server(port)
        sys.exit(0)

    # Check for dashboard mode
    if "--dashboard" in sys.argv or "--html" in sys.argv:
        create_dashboard_html(symbols)
        sys.exit(0)

    # Check for pipeline creation
    if "--pipeline" in sys.argv:
        create_pipeline_script(symbols)
        sys.exit(0)

    # Check for CoPaw skill creation
    if "--skill" in sys.argv:
        create_copaw_skill()
        sys.exit(0)

    # Check for webhook test
    if "--webhook" in sys.argv:
        config = load_config()
        webhook_url = ""
        if "--url" in sys.argv:
            idx = sys.argv.index("--url") + 1
            if idx < len(sys.argv):
                webhook_url = sys.argv[idx]
        elif config.get('notifications', {}).get('webhook_url'):
            webhook_url = config['notifications']['webhook_url']

        if webhook_url:
            send_webhook(webhook_url, {"test": True, "symbols": symbols})
        else:
            print("[Webhook] No webhook URL provided")
            print("Usage: --webhook --url https://your-webhook.com")
        sys.exit(0)

    # Check for notify mode (with webhook)
    if "--notify" in sys.argv:
        notify_alerts(symbols)
        sys.exit(0)

    # Check for cache clear
    if "--clear-cache" in sys.argv:
        clear_cache()
        sys.exit(0)

    # Check for config
    if "--config" in sys.argv:
        config = load_config()
        print("\n[Config]")
        print(json.dumps(config, indent=2))
        sys.exit(0)

    # Check for portfolio mode
    if "--portfolio" in sys.argv or "--portfolio-view" in sys.argv:
        show_portfolio()
        sys.exit(0)

    # Check for portfolio add
    if "--portfolio-add" in sys.argv:
        idx = sys.argv.index("--portfolio-add") + 1
        if idx + 2 <= len(sys.argv):
            sym = sys.argv[idx]
            shares = sys.argv[idx + 1]
            cost = sys.argv[idx + 2]
            add_position(sym, shares, cost)
        else:
            print("[Portfolio] Usage: --portfolio-add NVDA 100 170")
        sys.exit(0)

    # Check for portfolio remove
    if "--portfolio-remove" in sys.argv:
        idx = sys.argv.index("--portfolio-remove") + 1
        if idx < len(sys.argv):
            remove_position(sys.argv[idx])
        sys.exit(0)

    # Check for news
    if "--news" in sys.argv or "-n" in sys.argv:
        show_news(symbols)
        sys.exit(0)

    # Check for screener
    if "--screener" in sys.argv or "--screen" in sys.argv:
        min_score = 60
        min_upside = 15
        max_pe = 40

        if "--min-score" in sys.argv:
            idx = sys.argv.index("--min-score") + 1
            if idx < len(sys.argv):
                min_score = float(sys.argv[idx])

        if "--min-upside" in sys.argv:
            idx = sys.argv.index("--min-upside") + 1
            if idx < len(sys.argv):
                min_upside = float(sys.argv[idx])

        if "--max-pe" in sys.argv:
            idx = sys.argv.index("--max-pe") + 1
            if idx < len(sys.argv):
                max_pe = float(sys.argv[idx])

        show_screener(min_score, min_upside, max_pe)
        sys.exit(0)

    # Check for email report
    if "--email" in sys.argv or "--mail" in sys.argv:
        recipients = []
        if "--to" in sys.argv:
            idx = sys.argv.index("--to") + 1
            while idx < len(sys.argv) and not sys.argv[idx].startswith("-"):
                recipients.append(sys.argv[idx])
                idx += 1

        send_email_report(symbols, recipients if recipients else None)
        sys.exit(0)

    # Check for summary mode
    if "--summary" in sys.argv or "-s" in sys.argv:
        print("\n[ Stock Summary ]\n")
        results = []
        for sym in symbols:
            q = fetch(sym)
            if q.get("success"):
                fin = FD.get(sym, FD["AAPL"])
                analyst = ANALYST.get(sym, {"target": q["price"], "rating": "N/A"})
                dcf_scenarios = calc_dcf_scenarios(q, fin)
                dcf_base = dcf_scenarios["Base"]
                target = analyst["target"] * 0.6 + dcf_base * 0.4
                upside = (target - q["price"]) / q["price"] * 100

                quality = calc_quality_score(q, fin)
                value = calc_value_score(q, fin, {"percentile": 50})
                momentum_s = calc_momentum_score(q, {"range_pos": 50, "momentum": "Neutral"})
                growth = calc_growth_score(q, fin)
                sc = int(quality * 0.30 + value * 0.25 + momentum_s * 0.20 + growth * 0.25)

                results.append({
                    "symbol": sym,
                    "price": q["price"],
                    "target": target,
                    "upside": upside,
                    "score": sc,
                    "pe": q["pe"],
                    "peg": q["peg"],
                    "fcf_yield": fin.get("fcf_yield", 0),
                    "div_yield": fin.get("div_yield", 0),
                    "roe": fin["roe"],
                    "beta": q["beta"],
                    "rating": rating(upside),
                })

        results.sort(key=lambda x: x["upside"], reverse=True)
        for r in results:
            print_summary_card(r)
        sys.exit(0)

    # Default: single stock report
    sym = sys.argv[1].upper() if len(sys.argv) > 1 else "META"

    # Check for Chinese flag
    if "--cn" in sys.argv or "-cn" in sys.argv:
        generate_report_cn(sym)
    else:
        generate_report(sym)

    print("\n" + "="*60)
    print("Quick Commands:")
    print("   --cn        Chinese report")
    print("   --compare   Multi-stock comparison")
    print("   --watch     Watch mode (auto-refresh)")
    print("   --json      JSON output")
    print("   --summary   Summary mode")
    print("   --csv       Export to CSV")
    print("   --xlsx      Export to Excel")
    print("   --chart     Generate chart")
    print("   --db        Save to database")
    print("   --alert     Price alerts")
    print("   --obsidian  Sync to Obsidian")
    print("   --cron      Setup cron job")
    print("   --api       Start API server")
    print("   --dashboard Create HTML dashboard")
    print("   --pipeline  Create batch pipeline")
    print("   --skill     Create CoPaw skill")
    print("   --notify    Send notifications")
    print("   --help      Show all options")
    print("="*60)